# -*- coding: utf-8 -*-
"""
core/io.py

IO documental canónico:
- indexa enlaces KIID en el Excel maestro (una sola pasada)
- descarga PDFs con retries
- extrae texto (pdfplumber + OCR opcional)
- devuelve texto y metadata estructurada (incluye KIID_PDF_BYTES cuando hay texto)
"""

from typing import Optional, Dict, List, Tuple
from io import BytesIO
import os
import hashlib
import datetime
import gc

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

import pdfplumber
try:
    import pytesseract
    _HAS_TESSERACT = True
except Exception:
    _HAS_TESSERACT = False

from openpyxl import load_workbook

# Config
HEADERS = {"User-Agent": "Mozilla/5.0"}
REQUEST_TIMEOUT  = 15     # timeout por intento en segundos
REQUEST_RETRIES  = 3     # reintentos tras fallo (4 intentos totales)
# Demoras entre reintentos: backoff_factor=1 → 1s, 2s, 4s (formula: factor × 2^(n-1))

OCR_ENABLED = True
OCR_LANG = "spa+eng"
OCR_DPI = 300

# ── BL-DLA-1 kill-switch ──────────────────────────────────────────────────────
# DLA_ENABLED = True  activa la serialización 2D-aware (Sub-fase 1D: despliegue global).
# DLA_ENABLED = False revierte al comportamiento original sin tocar más código.
# Para roll-back inmediato: cambiar a False y reiniciar el pipeline.
DLA_ENABLED = True   # Sub-fase 1B: desactivado por defecto hasta validar piloto 1C.
# ─────────────────────────────────────────────────────────────────────────────

# ── BL-DLA-2 kill-switch ──────────────────────────────────────────────────────
# DLA_TABLE_SERIALIZATION_ENABLED = True  activa extracción de tablas Cat.1+2
#     (Sub-fase 2B: desactivado por defecto hasta validar piloto 2C).
# DLA_TABLE_SERIALIZATION_ENABLED = False no altera el comportamiento actual.
# Kill-switch INDEPENDIENTE de DLA_ENABLED: ambos pueden combinarse libremente.
# Para roll-back inmediato: cambiar a False y reiniciar el pipeline.
DLA_TABLE_SERIALIZATION_ENABLED = True  # Sub-fase 2B: desactivado hasta piloto 2C.
# ─────────────────────────────────────────────────────────────────────────────

MAX_PDF_PAGES = 3        # KIID = 2–3 páginas

KIID_CACHE_DAYS = 180    # días antes de re-verificar el PDF (Opción A)

MAX_PDF_MB = 20
MAX_PDF_BYTES = MAX_PDF_MB * 1024 * 1024

# Directorio de almacenamiento físico de PDFs KIID (BL-KIID-STORE).
# Cada fichero se guarda como {ISIN}.pdf. Al actualizar un KIID, el fichero
# antiguo se versiona con sufijo _YYYYMMDD (fecha de creación del original).
KIID_STORAGE_DIR = r"C:\data\fondos\kiid"

# ── BL-KIID-STORE kill-switch ─────────────────────────────────────────────────
# KIID_PHYSICAL_STORAGE_ENABLED = True  activa el guardado físico del PDF en disco
#     (store_kiid_pdf) desde get_kiid_for_isin cuando se descarga un PDF nuevo/mod.
# KIID_PHYSICAL_STORAGE_ENABLED = False no escribe nada en disco — comportamiento
#     idéntico al anterior a esta funcionalidad.
# Independiente de los kill-switches DLA. Roll-back inmediato: cambiar a False.
KIID_PHYSICAL_STORAGE_ENABLED = True
# ─────────────────────────────────────────────────────────────────────────────

# ── BL-KIID-LOCAL-FIRST kill-switch ──────────────────────────────────────────
# KIID_LOCAL_FIRST_ENABLED = True  activa la estrategia de carga híbrida:
#     antes de descargar por URL, get_kiid_for_isin busca el PDF en el
#     repositorio local KIID_STORAGE_DIR ({ISIN}.pdf). Solo actúa cuando la
#     caché A de texto en BD NO acierta (es decir, en el camino de re-extracción).
# KIID_LOCAL_FIRST_ENABLED = False comportamiento idéntico al anterior:
#     caché A (texto BD) → descarga directa por URL del Excel maestro.
# Este flag define el valor por DEFECTO del modo 'auto' (ver kiid_source en
# get_kiid_for_isin). El parámetro kiid_source puede forzar la modalidad por
# ejecución y tiene prioridad sobre este flag cuando es != 'auto'.
# Independiente de los demás kill-switches. Roll-back inmediato: cambiar a False.
KIID_LOCAL_FIRST_ENABLED = True
# ─────────────────────────────────────────────────────────────────────────────


# Cache ligero: excel_path -> { ISIN: [url, ...] }
_ISIN_LINKS_INDEX: Dict[str, Dict[str, List[str]]] = {}


def _requests_session() -> requests.Session:
    s = requests.Session()
    retries = Retry(
        total            = REQUEST_RETRIES,
        backoff_factor   = 1,                          # esperas: 1s, 2s, 4s
        status_forcelist = (500, 502, 503, 504),       # errores de servidor
        raise_on_status  = False,
    )
    s.mount("https://", HTTPAdapter(max_retries=retries))
    s.mount("http://", HTTPAdapter(max_retries=retries))
    s.headers.update(HEADERS)
    return s


def extract_text_from_pdf_bytes(
    pdf_bytes: bytes,
    ocr_enabled: bool = OCR_ENABLED,
    ocr_lang: str = OCR_LANG,
    ocr_dpi: int = OCR_DPI,
) -> str:
    # ── BL-DLA-1: despacho 2D-aware ──────────────────────────────────────────
    # Si DLA_ENABLED, delegar en dla_extractor.extract_text_dla_aware().
    # Si falla (fallback=True) o DLA_ENABLED=False, ejecutar lógica original.
    if DLA_ENABLED:
        try:
            from core.dla_extractor import extract_text_dla_aware, emit_dla_log
            dla_text, dla_meta = extract_text_dla_aware(
                pdf_bytes,
                ocr_enabled=ocr_enabled,
                ocr_lang=ocr_lang,
                ocr_dpi=ocr_dpi,
            )
            if not dla_meta.get("fallback") and dla_text and dla_text.strip():
                emit_dla_log("(unknown)", dla_meta)   # ISIN no disponible en esta capa
                return dla_text
            # fallback: DLA falló o produjo texto vacío → continuar a lógica original
        except Exception:
            pass   # Cualquier error de importación o ejecución → lógica original
    # ── Lógica original (pdfplumber) — intacta ────────────────────────────────
    text_parts = []

    with BytesIO(pdf_bytes) as b:
        with pdfplumber.open(b) as pdf:
            for i, page in enumerate(pdf.pages):
                if i >= MAX_PDF_PAGES:
                    break

                try:
                    t = page.extract_text()
                except Exception:
                    t = None

                if t and t.strip():
                    text_parts.append(t)
                    continue

                if ocr_enabled and _HAS_TESSERACT:
                    try:
                        img = page.to_image(resolution=ocr_dpi).original
                        text_parts.append(
                            pytesseract.image_to_string(img, lang=ocr_lang)
                        )
                    except Exception:
                        continue

    return "\n".join(text_parts)



def pdf_sha256(pdf_bytes: bytes) -> str:
    return hashlib.sha256(pdf_bytes).hexdigest()


def _sha256_of_file(path: str, chunk_size: int = 65536) -> str:
    """
    Calcula el SHA-256 de un fichero en disco leyéndolo por bloques
    (no carga el fichero completo en memoria).

    Reutiliza el mismo algoritmo que pdf_sha256() — la diferencia es solo
    la fuente (disco vs. bytes ya en memoria). Para un PDF KIID (<20 MB)
    el coste es despreciable, pero el streaming evita picos de RAM si el
    almacén creciera o se reutilizara la función para otros documentos.
    """
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(chunk_size), b""):
            h.update(chunk)
    return h.hexdigest()


def store_kiid_pdf(
    pdf_bytes: bytes,
    isin: str,
    storage_dir: str = KIID_STORAGE_DIR,
) -> Dict:
    """
    Almacena físicamente un PDF KIID recién descargado en `storage_dir`,
    gestionando deduplicación por hash y versionado del fichero anterior.

    Nombre del fichero: {ISIN}.pdf  (el ISIN es el identificador estable
    y único del fondo; las URLs de KIID no siempre tienen nombre fiable).

    Lógica:
      1. Crea el directorio de destino si no existe.
      2. Si NO existe un fichero previo para el ISIN → lo escribe directamente.
      3. Si SÍ existe:
         - Calcula SHA-256 del fichero local y del PDF descargado.
         - Hashes iguales  → descarta la descarga, no escribe nada (idéntico).
         - Hashes distintos → versiona el antiguo añadiendo _YYYYMMDD (fecha de
           CREACIÓN del fichero original, vía os.path.getctime) justo antes de
           la extensión, y luego escribe el nuevo PDF con el nombre original.

    Parámetros:
        pdf_bytes   — bytes del PDF recién descargado (ya en memoria).
        isin        — ISIN del fondo; determina el nombre del fichero.
        storage_dir — directorio de destino (por defecto KIID_STORAGE_DIR).

    Devuelve un dict con el resultado de la operación:
        {
            "action":        "WRITTEN_NEW" | "UNCHANGED" | "UPDATED" | "ERROR",
            "path":          ruta final del PDF (o None si ERROR),
            "archived_path": ruta del fichero versionado (solo en UPDATED),
            "hash":          SHA-256 del PDF almacenado,
            "error":         detalle técnico (solo si action == "ERROR"),
        }

    Notas de plataforma:
        - os.path.getctime() en Windows devuelve la fecha de CREACIÓN real del
          fichero, que es exactamente lo que pide el requisito ("fecha en la que
          ese archivo original fue depositado en el sistema"). En Linux esa misma
          llamada devuelve el último cambio de metadatos del inodo; si el almacén
          migrara a Linux habría que sustituirla por st_birthtime o un registro
          propio de la fecha de depósito.
        - La función NUNCA lanza excepción hacia el llamante: cualquier fallo se
          captura y se reporta vía action="ERROR" + campo "error", para no
          degradar el pipeline de extracción (consistente con el patrón DLA-2).
    """
    result = {
        "action": None,
        "path": None,
        "archived_path": None,
        "hash": None,
        "error": None,
    }

    try:
        if not pdf_bytes:
            result["action"] = "ERROR"
            result["error"] = "empty_pdf_bytes"
            return result

        if not isin or not str(isin).strip():
            result["action"] = "ERROR"
            result["error"] = "empty_isin"
            return result

        isin = str(isin).strip()

        # 1. Crear el directorio de destino si no existe.
        os.makedirs(storage_dir, exist_ok=True)

        filename = f"{isin}.pdf"
        dest_path = os.path.join(storage_dir, filename)

        new_hash = pdf_sha256(pdf_bytes)
        result["hash"] = new_hash

        # 2. No existe fichero previo → escribir directamente.
        if not os.path.exists(dest_path):
            with open(dest_path, "wb") as f:
                f.write(pdf_bytes)
            result["action"] = "WRITTEN_NEW"
            result["path"] = dest_path
            return result

        # 3. Existe fichero previo → comparar hashes.
        existing_hash = _sha256_of_file(dest_path)

        if existing_hash == new_hash:
            # Idéntico → descartar la descarga, no escribir nada.
            result["action"] = "UNCHANGED"
            result["path"] = dest_path
            return result

        # Distinto → versionar el antiguo con su fecha de creación.
        try:
            ctime = os.path.getctime(dest_path)
            date_suffix = datetime.datetime.fromtimestamp(ctime).strftime("%Y%m%d")
        except Exception:
            # Fallback defensivo: si no se puede leer la fecha de creación,
            # usar la fecha actual para no perder el versionado.
            date_suffix = datetime.datetime.now().strftime("%Y%m%d")

        root, ext = os.path.splitext(filename)        # ("ES0000000000", ".pdf")
        archived_name = f"{root}_{date_suffix}{ext}"   # "ES0000000000_20260524.pdf"
        archived_path = os.path.join(storage_dir, archived_name)

        # Versionado: un único fichero por fecha ({ISIN}_YYYYMMDD.pdf).
        # Se usa os.replace (no os.rename) porque es atómico y multiplataforma:
        # si ya existiera el archivado de esa misma fecha — caso raro, p.ej. el
        # pipeline reprocesa el fondo dos veces el mismo día con PDFs distintos —
        # lo sobrescribe en silencio, mientras que os.rename lanzaría
        # FileExistsError en Windows. Así el nombre queda siempre sin sufijos _N.
        os.replace(dest_path, archived_path)

        # Escribir el nuevo PDF con el nombre original.
        with open(dest_path, "wb") as f:
            f.write(pdf_bytes)

        result["action"] = "UPDATED"
        result["path"] = dest_path
        result["archived_path"] = archived_path
        return result

    except Exception as e:
        result["action"] = "ERROR"
        result["error"] = str(e)
        return result


def _load_local_kiid(
    isin: str,
    storage_dir: str = KIID_STORAGE_DIR,
) -> Optional[bytes]:
    """
    Lee el PDF KIID de un ISIN desde el repositorio local ({ISIN}.pdf).

    Devuelve los bytes del fichero si existe y es legible, o None en cualquier
    otro caso (no existe, vacío, error de E/S o permisos). NUNCA lanza
    excepción: un fallo de lectura local debe degradar limpiamente al Flujo B
    (descarga por URL), nunca abortar get_kiid_for_isin.

    El repositorio se autoalimenta vía store_kiid_pdf: cada PDF descargado por
    el Flujo B queda como {ISIN}.pdf, por lo que en ciclos posteriores este
    helper lo encontrará para el Flujo A.
    """
    try:
        if not isin or not str(isin).strip():
            return None
        path = os.path.join(storage_dir, f"{str(isin).strip()}.pdf")
        if not os.path.isfile(path):
            return None
        with open(path, "rb") as f:
            data = f.read()
        if not data:           # fichero de 0 bytes → tratar como ausente
            return None
        return data
    except Exception:
        # Cualquier fallo de lectura → None → fallback a Flujo B.
        return None



# ============================================================
# TRANSICIÓN AUTOMÁTICA A FORCE_REFRESH POR ANTIGÜEDAD
# ============================================================

def mark_stale_for_refresh(
    conn,
    max_age_days: int = KIID_CACHE_DAYS,
    max_funds: int = 100,
) -> int:
    """
    Marca como FORCE_REFRESH los fondos cuyo KIID no se ha re-descargado
    en más de max_age_days días.

    Parámetros:
        conn         — conexión SQLite activa
        max_age_days — antigüedad máxima en días (por defecto KIID_CACHE_DAYS=180)
        max_funds    — máximo de fondos a marcar por ejecución (evita avalanchas)

    Devuelve el número de fondos marcados.

    Llamar UNA VEZ al inicio del ciclo, antes de lanzar los bloques:
        from core.io import mark_stale_for_refresh
        n = mark_stale_for_refresh(conn, max_age_days=180, max_funds=50)
        print(f"[INFO] {n} fondos marcados FORCE_REFRESH por antigüedad")

    Con max_funds=50 y un universo de 3.204 fondos, el refresh completo
    se distribuye en ~64 ciclos (180 días / 50 fondos/ciclo ≈ 3,6 fondos/día
    de media), sin generar avalanchas de requests al servidor.
    """
    cutoff = (
        datetime.datetime.utcnow() - datetime.timedelta(days=max_age_days)
    ).isoformat(timespec="seconds")

    # Seleccionar los más antiguos primero (FIFO de antigüedad)
    sql = """
        UPDATE fund_kiid_metadata
        SET    KIID_Status = 'FORCE_REFRESH'
        WHERE  ISIN IN (
            SELECT ISIN
            FROM   fund_kiid_metadata
            WHERE  KIID_Status   IN ('OK', 'CACHED')
            AND    KIID_Class     = 1
            AND    (KIID_Downloaded_At IS NULL
                    OR KIID_Downloaded_At < ?)
            ORDER  BY KIID_Downloaded_At ASC NULLS FIRST
            LIMIT  ?
        )
        AND    KIID_Class = 1;
    """
    cursor = conn.execute(sql, (cutoff, max_funds))
    conn.commit()
    return cursor.rowcount


def _build_isin_links_index(excel_path: str) -> Dict[str, List[str]]:
    """
    Abre el workbook (modo normal) y construye un dict ISIN -> [URL...].
    Guarda el índice en _ISIN_LINKS_INDEX[excel_path].
    """
    excel_path = str(excel_path)
    if excel_path in _ISIN_LINKS_INDEX:
        return _ISIN_LINKS_INDEX[excel_path]

    index: Dict[str, List[str]] = {}
    wb = load_workbook(excel_path, data_only=True, read_only=False)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2):
                try:
                    if len(row) < 3:
                        continue
                    cell_isin = row[2]
                    if cell_isin is None or cell_isin.value is None:
                        continue
                    isin = str(cell_isin.value).strip()
                    if not isin:
                        continue

                    urls = index.setdefault(isin, [])
                    cell_link = row[3] if len(row) > 3 else None

                    if cell_link is not None:
                        link_obj = getattr(cell_link, "hyperlink", None)
                        if link_obj and getattr(link_obj, "target", None):
                            url = getattr(link_obj, "target")
                            if isinstance(url, str) and url.strip():
                                urls.append(url.strip())
                                continue
                        val = cell_link.value
                        if isinstance(val, str) and "http" in val.lower():
                            urls.append(val.strip())
                            continue
                except Exception:
                    continue
    finally:
        try:
            wb.close()
        except Exception:
            pass

    # normalizar: dedupe y ordenar por heurística (determinismo)
    for k, v in list(index.items()):
        seen = []
        for u in v:
            if u not in seen:
                seen.append(u)
        def _score(u):
            score = 0
            if isinstance(u, str) and u.lower().endswith(".pdf"):
                score -= 10
            if isinstance(u, str) and "kiid" in u.lower():
                score -= 5
            return score
        seen.sort(key=lambda x: (_score(x), x))
        index[k] = seen

    _ISIN_LINKS_INDEX[excel_path] = index
    gc.collect()
    return index


def find_kiid_links_from_excel(isin: str, excel_path: str) -> List[str]:
    """
    Devuelve lista (posible vacía) de URLs para un ISIN usando el índice.
    """
    index = _build_isin_links_index(excel_path)
    return index.get(isin, [])


def download_pdf(url: str, session: Optional[requests.Session] = None) -> bytes:
    session = session or _requests_session()
    resp = session.get(url, timeout=REQUEST_TIMEOUT, stream=True)
    resp.raise_for_status()

    content = bytearray()
    for chunk in resp.iter_content(chunk_size=8192):
        content.extend(chunk)
        if len(content) > MAX_PDF_BYTES:
            raise ValueError("pdf_too_large")

    return bytes(content)


def _process_pdf_bytes(
    pdf_bytes: bytes,
    isin: str,
    new_hash: str,
    source: str,
    url: Optional[str] = None,
) -> Tuple[Optional[str], Dict]:
    """
    Procesa por completo un PDF KIID ya en memoria, con independencia de su
    origen (repositorio local o descarga por URL). Centraliza la lógica de
    negocio común a ambos caminos (Principio #2 DRY): extracción de texto,
    serialización de tablas DLA-2, almacenamiento físico y construcción del
    bloque de metadatos del documento procesado.

    Esta función se extrajo del cuerpo del bucle `for url in links` de
    get_kiid_for_isin para que el Flujo A (local) y el Flujo B (remoto)
    apliquen EXACTAMENTE la misma lógica, en lugar de duplicarla.

    Args:
        pdf_bytes — bytes del PDF (de disco local o recién descargado).
        isin      — ISIN del fondo.
        new_hash  — SHA-256 ya calculado de pdf_bytes (se pasa para no recalcular).
        source    — 'LOCAL' o 'REMOTE'; se refleja en meta['KIID_Source'] (traza).
        url       — URL de origen si source == 'REMOTE'; None si 'LOCAL'.

    Returns:
        (kiid_text, meta_updates):
          - Si el texto es válido: (texto, dict con claves a fusionar en meta y
            KIID_Status='OK').
          - Si la extracción falla o el texto es vacío: (None, dict con
            KIID_Status de error). El llamante decide si continuar (otro enlace)
            o degradar.
    """
    meta_updates: Dict = {}

    # 1. Extracción de texto del binario.
    try:
        kiid_text = extract_text_from_pdf_bytes(pdf_bytes)
    except Exception as e:
        meta_updates["KIID_Status"] = "DOWNLOAD_ERROR"
        meta_updates["KIID_Error"] = f"text_extraction_error: {e}"
        meta_updates["KIID_Source"] = source
        return None, meta_updates

    if not kiid_text or not kiid_text.strip():
        meta_updates["KIID_Status"] = "EMPTY_TEXT"
        meta_updates["KIID_Error"] = f"empty_text_from_{source.lower()}"
        meta_updates["KIID_Source"] = source
        return None, meta_updates

    # 2. BL-DLA-2: extraer tablas Cat.1+2 y enriquecer el texto.
    #    Idéntico para local y remoto. Nunca degrada kiid_text ante fallo.
    #
    # FIX-DATA-INTEGRITY-2 (2026-06-17): la línea original mutaba kiid_text
    # IN-PLACE con el bloque DLA2 recién serializado. kiid_text es el valor
    # que esta función devuelve y que termina en Raw_KIID_Text (vía
    # get_kiid_for_isin -> pipeline.py L1750 -> sqlite_writer COALESCE),
    # horneando PERMANENTEMENTE contenido derivado (grid '|||') dentro de lo
    # que debería ser texto puro extraído del PDF. Esta es la fuente ORIGINAL
    # de la corrupción confirmada en 30 fondos (FIX-DATA-INTEGRITY-1 en
    # get_kiid_for_isin solo arregla la COMPOUNDING posterior en ciclos
    # cacheados; esta es la causa raíz en el momento de la extracción fresca).
    # Fix: kiid_text permanece puro. El texto enriquecido para
    # extract_priips_costs/extract_ucits_costs se expone vía
    # meta_updates["Fed_Text_For_Cost"] (no persistido).
    _dla2_table_text = None
    _fed_text_for_cost = None
    if DLA_TABLE_SERIALIZATION_ENABLED:
        try:
            from core.dla_table_serializer import (
                serialize_tables, emit_table_log,
            )
            # BL-COST DLA2 fix: pasar el texto YA OCR'd (Raw) para cubrir PDFs
            # escaneados y de columnas entrelazadas (el grid solo cubre PDFs con tabla).
            _t2_text, _t2_meta = serialize_tables(pdf_bytes, text=kiid_text, debug=False)
            if _t2_text and _t2_text.strip():
                _dla2_table_text = _t2_text
                _fed_text_for_cost = kiid_text + "\n" + _t2_text
                emit_table_log(isin, _t2_meta)
        except Exception:
            pass   # Nunca degradar kiid_text por un fallo de DLA-2

    # 3. BL-KIID-STORE: persistir físicamente el PDF en disco.
    #    En el Flujo A (LOCAL) el binario YA está en disco; aun así se invoca
    #    store_kiid_pdf por coherencia: si el contenido leído coincide con el
    #    almacenado dará UNCHANGED (no reescribe), y si por algún motivo difiere
    #    versiona y actualiza. El coste es un hash de fichero, despreciable.
    if KIID_PHYSICAL_STORAGE_ENABLED:
        try:
            _store_result = store_kiid_pdf(pdf_bytes, isin)
            meta_updates["KIID_Stored_Action"] = _store_result.get("action")
            meta_updates["KIID_Stored_Path"]   = _store_result.get("path")
            if _store_result.get("archived_path"):
                meta_updates["KIID_Archived_Path"] = _store_result.get("archived_path")
        except Exception as e:
            meta_updates["KIID_Stored_Action"] = "ERROR"
            meta_updates["KIID_Stored_Path"]   = None
            meta_updates["KIID_Store_Error"]   = str(e)

    # 4. Metadatos del documento procesado (KIID válido).
    meta_updates.update({
        "KIID_URL":            url,
        "KIID_PDF_Hash":       new_hash,
        "KIID_Downloaded_At":  datetime.datetime.utcnow().isoformat(timespec="seconds"),
        "KIID_PDF_BYTES":      pdf_bytes,          # temporal (pipeline)
        "KIID_Status":         "OK",
        "KIID_Error":          None,
        "KIID_Source":         source,             # traza: LOCAL | REMOTE
        "DLA2_Table_Text":     _dla2_table_text,   # None si flag desactivado
        # FIX-DATA-INTEGRITY-2: texto enriquecido SOLO para
        # extract_priips_costs/extract_ucits_costs en este ciclo.
        # Nunca usar para Raw_KIID_Text.
        "Fed_Text_For_Cost":   _fed_text_for_cost,
    })

    return kiid_text, meta_updates


def get_kiid_for_isin(
    isin: str,
    excel_master_path: Optional[str],
    conn=None,
    kiid_source: str = "auto",
) -> Tuple[Optional[str], Dict]:
    """
    Intentar localizar y extraer el texto del KIID para un ISIN.

    Devuelve:
        (kiid_text o None, metadata dict)

    metadata incluye:
        - KIID_Status  (estado documental: OK / CACHED / DOWNLOAD_ERROR / ...)
        - KIID_Error   (detalle técnico)
        - KIID_PDF_BYTES (solo si KIID_Status == OK — documento procesado)
        - KIID_Source  (traza del origen: CACHE / LOCAL / REMOTE)

    Orden de resolución (BL-KIID-LOCAL-FIRST):
        1. Caché A: si conn tiene Raw_KIID_Text para el ISIN (KIID_Status
           OK/CACHED) → devuelve el texto cacheado SIN tocar disco ni red.
           Es el camino del ~98% del corpus en régimen estacionario.
        2. Solo si la caché A no acierta, y según kiid_source:
           - Flujo A (local): leer {ISIN}.pdf del repositorio local. Si el hash
             coincide con BD → reusar texto BD; si difiere → procesar completo.
           - Flujo B (remoto): descargar por URL del Excel maestro y procesar.
             El PDF descargado queda en local (store_kiid_pdf) para ciclos futuros.

    Args:
        kiid_source — modalidad de carga del binario cuando la caché A falla:
            'auto'   (por defecto): local-first si KIID_LOCAL_FIRST_ENABLED;
                     en caso contrario, directo a remoto. Si el local no tiene
                     el fichero, hace fallback a remoto.
            'local'  : fuerza local-first (con fallback a remoto si no existe
                     el fichero o falla su lectura).
            'remote' : fuerza descarga por URL; ignora el repositorio local.
    """

    meta = {
        "ISIN": isin,
        "KIID_URL": None,
        "KIID_PDF_Hash": None,
        "KIID_Downloaded_At": None,
        "KIID_Status": None,
        "KIID_Error": None,
    }

    # -------------------------------------------------
    # Lectura de datos previos en BD (texto, hash, url, tabla DLA-2).
    #
    # Causa raíz del fix: esta lectura NO debe filtrar por KIID_Status. El hash
    # previo (KIID_PDF_Hash) es necesario para el reúso por hash de los Flujos
    # A (local) y B (remoto) — que evita re-OCR/SRRI visual cuando el PDF no ha
    # cambiado — y ese camino se recorre precisamente cuando el estado es
    # FORCE_REFRESH (estado que la caché A excluye por diseño). Antes, el hash
    # solo se leía dentro de la rama de caché A que retorna de inmediato, por lo
    # que _cached_hash siempre era None al llegar a los flujos de descarga y las
    # ramas de reúso por hash eran inalcanzables (código muerto).
    #
    # Solución: leer SIEMPRE los datos previos (incluido KIID_Status), y separar
    # dos responsabilidades:
    #   1. Caché A (retorno temprano): solo si estado ∈ (OK, CACHED) y hay texto.
    #   2. Datos previos (_cached_*): disponibles para los Flujos A/B aunque el
    #      estado sea FORCE_REFRESH, habilitando el reúso por hash.
    # -------------------------------------------------
    _cached_text = _cached_hash = _cached_url = None
    _cached_table_text = None
    _cached_downloaded_at = None
    _cached_status = None
    _fed_text_for_cost = None   # FIX-DATA-INTEGRITY-1: always defined, any branch
    if conn is not None:
        try:
            cached = conn.execute(
                """SELECT Raw_KIID_Text, KIID_PDF_Hash, KIID_URL,
                          KIID_Downloaded_At, DLA2_Table_Text, KIID_Status
                   FROM fund_kiid_metadata
                   WHERE ISIN = ? AND KIID_Class = 1
                     AND Raw_KIID_Text IS NOT NULL""",
                (isin,)
            ).fetchone()
            if cached and cached[0] and cached[0].strip():
                _cached_text          = cached[0]
                _cached_hash          = cached[1]
                _cached_url           = cached[2]
                _cached_downloaded_at = cached[3]
                _cached_table_text    = cached[4]   # DLA2_Table_Text — puede ser NULL
                _cached_status        = cached[5]

                # ── BL-DLA-2: enriquecer texto cacheado con tablas Cat.1+2 ──
                # Si DLA_TABLE_SERIALIZATION_ENABLED y ya hay texto de tablas en BD,
                # concatenar — idéntico a cómo se hace en la descarga real.
                # Si DLA2_Table_Text es NULL (fondo no procesado aún con DLA-2),
                # usar Raw_KIID_Text sin modificar (no degrade).
                #
                # FIX-DATA-INTEGRITY-1 (2026-06-17): la línea original mutaba
                # _cached_text IN-PLACE. _cached_text es el mismo valor que esta
                # función devuelve como kiid_text, y pipeline.py L1750 lo asigna
                # DIRECTAMENTE a kiid_record["Raw_KIID_Text"] -> sqlite_writer's
                # COALESCE lo persiste de vuelta en la columna Raw_KIID_Text en
                # CADA ciclo (CACHED o FORCE_REFRESH-con-hash-reuse) en que este
                # bloque se ejecuta. Resultado: el bloque DLA2 (potencialmente
                # obsoleto/incorrecto, generado por una versión anterior del
                # serializador) queda incrustado PERMANENTEMENTE en Raw_KIID_Text,
                # se relee en el siguiente ciclo, y se reconcatena de nuevo
                # (corrupción compuesta — confirmado en 30 fondos, todos con
                # KIID_Status=FORCE_REFRESH y hash de PDF sin cambios, por lo que
                # el atajo de reúso por hash nunca vuelve a re-extraer del PDF).
                # Fix: NUNCA mutar _cached_text. El texto enriquecido para uso
                # exclusivo de extract_priips_costs/extract_ucits_costs se expone
                # vía meta["Fed_Text_For_Cost"] (variable NUEVA, no persistida).
                # Raw_KIID_Text permanece puro en todos los caminos de retorno.
                _fed_text_for_cost = None
                if (DLA_TABLE_SERIALIZATION_ENABLED
                        and _cached_table_text
                        and _cached_table_text.strip()):
                    _fed_text_for_cost = _cached_text + "\n" + _cached_table_text

                # ── Caché A (retorno temprano) ──────────────────────────────
                # Solo si el estado permite servir desde caché. FORCE_REFRESH
                # NO entra aquí: cae a los flujos de obtención de binario, pero
                # conservando _cached_text/_cached_hash para el reúso por hash.
                if _cached_status in ('OK', 'CACHED'):
                    meta.update({
                        "KIID_URL":            _cached_url,
                        "KIID_PDF_Hash":       _cached_hash,
                        "KIID_Status":         "CACHED",
                        "KIID_Downloaded_At":  _cached_downloaded_at,
                        "KIID_Error":          None,
                        "KIID_Class":          1,
                        "KIID_Source":         "CACHE",
                        # FIX-DATA-INTEGRITY-1: texto enriquecido SOLO para
                        # extract_priips_costs/extract_ucits_costs en este
                        # ciclo. Nunca usar para Raw_KIID_Text.
                        "Fed_Text_For_Cost":   _fed_text_for_cost,
                    })
                    return _cached_text, meta
        except Exception:
            _cached_text = _cached_hash = _cached_url = None
            _cached_table_text = None
            _cached_status = None
            _fed_text_for_cost = None

    # =================================================================
    # La caché A no acertó (o no hay conn): hay que obtener el binario.
    # Resolver la modalidad efectiva (BL-KIID-LOCAL-FIRST).
    # =================================================================
    #   kiid_source == 'remote' → ignorar repositorio local (solo Flujo B).
    #   kiid_source == 'local'  → forzar local-first (con fallback a remoto).
    #   kiid_source == 'auto'   → local-first si KIID_LOCAL_FIRST_ENABLED.
    _src = (kiid_source or "auto").strip().lower()
    if _src not in ("auto", "local", "remote"):
        _src = "auto"   # valor desconocido → comportamiento seguro por defecto

    _try_local = (_src == "local") or (_src == "auto" and KIID_LOCAL_FIRST_ENABLED)

    # -------------------------------------------------
    # FLUJO A — Repositorio local ({ISIN}.pdf)
    # -------------------------------------------------
    # Solo se intenta si la caché A de texto NO acertó. Cualquier fallo de
    # lectura o procesamiento del binario local degrada limpiamente al Flujo B
    # (descarga remota): el local nunca empeora la robustez actual.
    if _try_local:
        local_bytes = _load_local_kiid(isin)
        if local_bytes is not None:
            try:
                local_hash = pdf_sha256(local_bytes)

                # Si el binario local coincide con el hash en BD, su texto ya
                # está cacheado: reusar (evita OCR + SRRI visual). Equivale a la
                # Caché B, alimentada desde disco en lugar de desde la red.
                # BL-CACHE-RESTORE (v20): el reúso rápido por hash se aplica
                # TAMBIÉN en FORCE_REFRESH. Si el binario local es byte-idéntico
                # al de BD, su texto/SRRI ya están cacheados: re-OCR sería puro
                # desperdicio (y a escala de corpus, una avalancha de Tesseract).
                # El fix BL-COST anterior descartaba el reúso para obtener
                # KIID_PDF_BYTES; aquí se logra lo mismo SIN reprocesar: se adjunta
                # el binario local a la meta, de modo que el bloque de coste del
                # pipeline (gated 'pdf_bytes is not None') corre sobre él sin
                # re-extraer el texto. Vía rápida restaurada + coste habilitado.
                if (_cached_text and _cached_hash and local_hash == _cached_hash):
                    meta.update({
                        "KIID_URL":            _cached_url,
                        "KIID_PDF_Hash":       local_hash,
                        "KIID_Downloaded_At":  datetime.datetime.utcnow().isoformat(timespec="seconds"),
                        "KIID_Status":         "OK",
                        "KIID_Error":          None,
                        "KIID_Class":          1,
                        "KIID_Source":         "LOCAL",
                        # DLA2_Table_Text no se incluye: el writer usa COALESCE
                        # y preservará el valor existente en BD (no hay re-extracción).
                        # FIX-DATA-INTEGRITY-1: idem Caché A — solo para coste,
                        # nunca para Raw_KIID_Text.
                        "Fed_Text_For_Cost":   _fed_text_for_cost,
                    })
                    # Adjuntar binario local: habilita el bloque de coste sin re-OCR.
                    meta["KIID_PDF_BYTES"] = local_bytes
                    return _cached_text, meta

                # Hash distinto (o sin texto/hash en BD, p.ej. PDF depositado
                # manualmente sin fila previa): procesar completo desde local.
                kiid_text, _updates = _process_pdf_bytes(
                    local_bytes, isin, local_hash, source="LOCAL", url=_cached_url
                )
                if kiid_text:
                    meta.update(_updates)
                    return kiid_text, meta
                # Texto vacío / extracción fallida en local → fallback a remoto.
            except Exception:
                # Fallo inesperado procesando el local → fallback a remoto.
                pass

    # Si la modalidad era estrictamente local pero no se pudo resolver,
    # igualmente caemos al Flujo B (fallback): preferimos un documento remoto
    # a no devolver nada. (El modo 'remote' nunca entra al bloque anterior.)

    # -------------------------------------------------
    # FLUJO B — Descarga remota por URL del Excel maestro (fallback)
    # -------------------------------------------------
    if not excel_master_path:
        meta["KIID_Status"] = "DOWNLOAD_ERROR"
        meta["KIID_Error"] = "no_excel_master_path"
        return None, meta

    try:
        links = find_kiid_links_from_excel(isin, excel_master_path)
    except Exception as e:
        meta["KIID_Status"] = "DOWNLOAD_ERROR"
        meta["KIID_Error"] = f"excel_index_error: {e}"
        return None, meta

    if not links:
        meta["KIID_Status"] = "NO_KIID"
        meta["KIID_Error"] = "no_links_found"
        return None, meta

    session = _requests_session()

    for url in links:
        try:
            pdf_bytes = download_pdf(url, session)

            # Caché B (remota): comparar hash. Si coincide con BD, reusar texto
            # (evita OCR + SRRI visual). Solo se llega aquí en FORCE_REFRESH.
            # BL-CACHE-RESTORE (v20): reúso rápido por hash también en
            # FORCE_REFRESH (binario sin cambios → sin re-OCR). Se adjunta el
            # binario ya descargado para habilitar el bloque de coste sin
            # reprocesar el texto. (Esta rama solo se alcanza si FLUJO A local
            # no resolvió; el coste de red ya se pagó en download_pdf.)
            new_hash = pdf_sha256(pdf_bytes)
            if (_cached_text and _cached_hash and new_hash == _cached_hash):
                meta.update({
                    "KIID_URL":            url,
                    "KIID_PDF_Hash":       new_hash,
                    "KIID_Downloaded_At":  datetime.datetime.utcnow().isoformat(timespec="seconds"),
                    "KIID_Status":         "OK",
                    "KIID_Error":          None,
                    "KIID_Class":          1,
                    "KIID_Source":         "REMOTE",
                    # DLA2_Table_Text no se incluye aquí: el writer usa COALESCE
                    # y preservará el valor existente en BD (no hay re-extracción).
                    # FIX-DATA-INTEGRITY-1: idem Caché A / Flujo A local.
                    "Fed_Text_For_Cost":   _fed_text_for_cost,
                })
                meta["KIID_PDF_BYTES"] = pdf_bytes
                return _cached_text, meta

            # PDF nuevo o modificado — procesar completo (lógica compartida).
            kiid_text, _updates = _process_pdf_bytes(
                pdf_bytes, isin, new_hash, source="REMOTE", url=url
            )
            if kiid_text:
                meta.update(_updates)
                return kiid_text, meta

            # Texto vacío / extracción fallida → registrar y probar otro enlace.
            meta.update({k: v for k, v in _updates.items()
                         if k in ("KIID_Status", "KIID_Error", "KIID_Source")})
            continue

        except ValueError as e:
            # Casos controlados (ej. pdf_too_large)
            if "pdf_too_large" in str(e):
                meta["KIID_Status"] = "PDF_TOO_LARGE"
                meta["KIID_Error"] = str(e)
            else:
                meta["KIID_Status"] = "DOWNLOAD_ERROR"
                meta["KIID_Error"] = str(e)
            continue

        except Exception as e:
            meta["KIID_Status"] = "DOWNLOAD_ERROR"
            meta["KIID_Error"] = str(e)
            continue

    # -------------------------------------------------
    # Ningún enlace válido
    # -------------------------------------------------
    if meta["KIID_Status"] is None:
       meta["KIID_Status"] = "NO_KIID"
       meta["KIID_Error"] = "no_valid_kiid_found"


    return None, meta
