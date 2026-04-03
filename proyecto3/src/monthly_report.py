# proyecto3/src/monthly_report.py
# -*- coding: utf-8 -*-
"""
Generador de informe mensual de la cartera.

Produce un Excel con las siguientes hojas:
  0_Portada         Resumen ejecutivo con regimen, semaforo y cartera
  1_Cartera         Detalle de fondos con pesos y metricas clave
  2_Regimen         Clasificacion macro actual e historica
  3_Backtesting     Resumen de resultados del backtesting
  4_Rotacion        Recomendaciones de rotacion si procede

Uso:
    from proyecto3.src.monthly_report import generate_report
    generate_report(conn, output_dir="c:/data/fondos/reports")
"""

import sqlite3
import pandas as pd
from pathlib import Path
from datetime import date
import sys

_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_ROOT))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl no instalado. Ejecuta: pip install openpyxl")

from proyecto3.src.regime_classifier import RegimeClassifier
from proyecto3.src.backtesting import Backtester
from shared.config import REPORTS_DIR as _DEFAULT_REPORTS_DIR


# ============================================================
# Estilos
# ============================================================

def _font(bold=False, size=11, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _center():
    return Alignment(horizontal="center", vertical="center")

TITLE_FONT  = _font(bold=True, size=13, color="FFFFFF")
TITLE_FILL  = _fill("1D3557")
HEADER_FONT = _font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = _fill("457B9D")

REGIME_COLORS = {
    "Expansion":              ("D4EDDA", "155724"),
    "Recalentamiento":        ("FFF3CD", "856404"),
    "Recalentamiento_Tardio": ("FFE5B4", "6B3A00"),
    "Estanflacion":           ("F8D7DA", "721C24"),
    "Contraccion":            ("D1ECF1", "0C5460"),
    "Shock_Energetico":       ("F8D7DA", "721C24"),
}

SEMAFORO_COLORS = {
    "Verde": ("D4EDDA", "155724"),
    "Ambar": ("FFF3CD", "856404"),
    "Rojo":  ("F8D7DA", "721C24"),
}


def _apply_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = _center()


def _autofit(ws, min_width=8, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ============================================================
# Calculo del objetivo minimo IPC + M3
# ============================================================

def _get_objetivo_ipc_m3(conn: sqlite3.Connection) -> dict:
    """
    Calcula el objetivo minimo anual de preservacion de patrimonio:
        objetivo = IPC_yoy + M3_yoy
    Devuelve dict con valores anuales y mensuales equivalentes.
    """
    # IPC ultimo dato disponible (media ES+EU)
    ipc_rows = conn.execute("""
        SELECT geography, value, date
        FROM series_macro
        WHERE indicator = 'ipc_index' AND geography IN ('ES','EU')
        ORDER BY date DESC
        LIMIT 4
    """).fetchall()

    ipc_yoy = None
    if ipc_rows:
        df_ipc = {}
        for geo, val, dt in ipc_rows:
            if geo not in df_ipc:
                df_ipc[geo] = (dt, float(val))

        # Calcular YoY aproximado desde los ultimos datos disponibles
        ipc_now = conn.execute("""
            SELECT geography, value FROM series_macro
            WHERE indicator='ipc_index' AND geography IN ('ES','EU')
            AND date = (SELECT MAX(date) FROM series_macro
                        WHERE indicator='ipc_index' AND geography='ES')
        """).fetchall()
        ipc_12m = conn.execute("""
            SELECT geography, value FROM series_macro
            WHERE indicator='ipc_index' AND geography IN ('ES','EU')
            AND date = (
                SELECT date FROM series_macro
                WHERE indicator='ipc_index' AND geography='ES'
                ORDER BY date DESC LIMIT 1 OFFSET 12
            )
        """).fetchall()

        if ipc_now and ipc_12m:
            now_map = {g: v for g, v in ipc_now}
            ago_map = {g: v for g, v in ipc_12m}
            vals = []
            for geo in ['ES', 'EU']:
                if geo in now_map and geo in ago_map and ago_map[geo] > 0:
                    vals.append((now_map[geo] - ago_map[geo]) / ago_map[geo])
            if vals:
                ipc_yoy = sum(vals) / len(vals)

    # M3 ultimo dato disponible
    m3_row = conn.execute("""
        SELECT value FROM series_macro
        WHERE indicator='m3_yoy'  AND geography='EU'
        ORDER BY date DESC LIMIT 1
    """).fetchone()
    m3_yoy = float(m3_row[0]) / 100 if m3_row else None

    objetivo_anual = None
    if ipc_yoy is not None and m3_yoy is not None:
        objetivo_anual = ipc_yoy + m3_yoy
    elif ipc_yoy is not None:
        objetivo_anual = ipc_yoy

    objetivo_mensual = ((1 + objetivo_anual) ** (1/12) - 1
                        if objetivo_anual is not None else None)

    return {
        "ipc_yoy":        ipc_yoy,
        "m3_yoy":         m3_yoy,
        "objetivo_anual": objetivo_anual,
        "objetivo_mensual": objetivo_mensual,
    }


# ============================================================
# Hojas del informe
# ============================================================

def _build_portada(ws, conn, clf, scenario_id="shock_energia_2026Q1"):
    """Hoja 0 — Portada con resumen ejecutivo."""
    ws.sheet_view.showGridLines = False

    # Titulo
    ws.merge_cells("A1:F1")
    ws["A1"] = f"INFORME MENSUAL DE CARTERA  |  {date.today().strftime('%B %Y').upper()}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 30

    # Regimen actual
    r = clf.classify_current()
    sem = clf.semaforo()
    reg_bg, reg_fg = REGIME_COLORS.get(r.regime, ("FFFFFF", "000000"))
    sem_bg, sem_fg = SEMAFORO_COLORS.get(sem.color, ("FFFFFF", "000000"))

    datos_regimen = [
        ("", ""),
        ("REGIMEN MACRO", r.regime),
        ("SEMAFORO", sem.color),
        ("IPC medio EU+ES", f"{r.ipc_yoy_avg*100:.1f}%" if r.ipc_yoy_avg else "N/D"),
        ("CLI Eurozona", f"{r.cli_eu:.1f}" if r.cli_eu else "N/D"),
        ("Tipo deposito BCE", f"{r.rate_deposit:.2f}%" if r.rate_deposit else "N/D"),
        ("Petroleo YoY", f"{r.oil_yoy*100:+.1f}%" if r.oil_yoy else "N/D"),
        ("", ""),
        ("PESOS DE CARTERA", ""),
        ("Defensiva", f"{r.weight_defensive:.0%}"),
        ("Equilibrada", f"{r.weight_balanced:.0%}"),
        ("Dinamica", f"{r.weight_dynamic:.0%}"),
    ]

    # Objetivo IPC+M3
    obj = _get_objetivo_ipc_m3(conn)
    datos_regimen += [
        ("", ""),
        ("OBJETIVO MINIMO ANUAL", ""),
        ("IPC medio EU+ES", f"{obj['ipc_yoy']*100:.1f}%" if obj['ipc_yoy'] else "N/D"),
        ("M3 Eurozona YoY", f"{obj['m3_yoy']*100:.1f}%" if obj['m3_yoy'] else "N/D"),
        ("Objetivo IPC+M3",
         f"{obj['objetivo_anual']*100:.1f}% anual / "
         f"{obj['objetivo_mensual']*100:.2f}% mensual"
         if obj['objetivo_anual'] else "N/D"),
    ]

    for i, (label, value) in enumerate(datos_regimen, 3):
        ws.cell(i, 1, label).font = _font(bold=True, size=10)
        cell_val = ws.cell(i, 2, value)
        cell_val.font = _font(size=10)
        if label == "REGIMEN MACRO":
            cell_val.fill = _fill(reg_bg)
            cell_val.font = _font(bold=True, size=10, color=reg_fg)
        elif label == "SEMAFORO":
            cell_val.fill = _fill(sem_bg)
            cell_val.font = _font(bold=True, size=10, color=sem_fg)

    # Senales del semaforo
    if sem.signals:
        row = 3 + len(datos_regimen) + 1
        ws.cell(row, 1, "SENALES:").font = _font(bold=True, size=10)
        for j, signal in enumerate(sem.signals):
            ws.cell(row + j + 1, 1, f"  ! {signal}").font = _font(size=10, color="721C24")

    _autofit(ws)


def _build_cartera(ws, conn, scenario_id="shock_energia_2026Q1"):
    """Hoja 1 — Detalle de fondos de la cartera actual."""
    ws.sheet_view.showGridLines = False

    ws["A1"] = f"CARTERA ACTUAL  |  Escenario: {scenario_id}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = _center()

    obj = _get_objetivo_ipc_m3(conn)
    obj_anual = obj["objetivo_anual"] * 100 if obj["objetivo_anual"] else None

    headers = ["Sub-cartera", "ISIN", "Nombre", "Naturaleza", "Gestora",
               "Peso %", "Score", "Retorno real %", "Max DD %",
               "Obj IPC+M3 %", "Exceso obj %", "Divisa contrib %",
               "Moneda", "Cobertura", "Geografia",
               "Beta Oil", "Beta Rate EU", "Beta Rate US",
               "Macro R2", "Alpha Persist.", "SRRI KIID", "SRRI Calc"]
    _apply_header(ws, headers, row=2)
    ws.merge_cells(f"A1:{get_column_letter(22)}1")

    rows = conn.execute("""
        SELECT pw.block, pw.isin, fm.Fund_Name, fm.Fund_Nature,
               fm.Management_Company,
               ROUND(pw.weight * 100, 1) as peso_master,
               ROUND(fs.score_total, 4) as score,
               ROUND(ret.value * 100, 2) as return_real,
               ROUND(dd.value * 100, 2)    as max_dd,
               ROUND(fx.value * 100, 2)    as fx_ann_pct,
               fm.Fund_Currency,
               fm.Hedging_Policy,
               fm.Geography,
               ROUND(oil.value, 4)          as beta_oil,
               ROUND(reu.value, 4)          as beta_rate_eu,
               ROUND(rus.value, 4)          as beta_rate_us,
               ROUND(r2.value, 3)           as macro_r2,
               ROUND(per.value, 3)           as alpha_persistence,
               fm.SRRI                         as srri_kiid,
               CAST(srri.value AS INTEGER)      as srri_calc
        FROM portfolio_weights pw
        JOIN fund_master fm ON fm.ISIN = pw.isin
        LEFT JOIN fund_scores fs ON fs.isin = pw.isin AND fs.block = pw.block
                                 AND fs.score_version = 'v1'
        LEFT JOIN fund_metrics ret ON ret.isin = pw.isin
                                   AND ret.metric = 'return_ann'
                                   AND ret.horizon = 'since_inception'
                                   AND ret.real_flag = 1
        LEFT JOIN fund_metrics dd  ON dd.isin = pw.isin
                                   AND dd.metric = 'max_drawdown'
                                   AND dd.horizon = 'since_inception'
                                   AND dd.real_flag = 0
        LEFT JOIN fund_metrics fx  ON fx.isin = pw.isin
                                   AND fx.metric = 'fx_contribution_ann'
                                   AND fx.horizon = 'since_inception'
                                   AND fx.real_flag = 0
        LEFT JOIN fund_metrics oil ON oil.isin = pw.isin
                                   AND oil.metric = 'beta_oil'
                                   AND oil.horizon = 'since_inception'
                                   AND oil.real_flag = 0
        LEFT JOIN fund_metrics reu ON reu.isin = pw.isin
                                   AND reu.metric = 'beta_rate_eu'
                                   AND reu.horizon = 'since_inception'
                                   AND reu.real_flag = 0
        LEFT JOIN fund_metrics rus ON rus.isin = pw.isin
                                   AND rus.metric = 'beta_rate_us'
                                   AND rus.horizon = 'since_inception'
                                   AND rus.real_flag = 0
        LEFT JOIN fund_metrics r2  ON r2.isin = pw.isin
                                   AND r2.metric = 'macro_r2'
                                   AND r2.horizon = 'since_inception'
                                   AND r2.real_flag = 0
        LEFT JOIN fund_metrics per ON per.isin = pw.isin
                                   AND per.metric = 'alpha_persistence'
                                   AND per.horizon = 'since_inception'
                                   AND per.real_flag = 0
        LEFT JOIN fund_metrics srri ON srri.isin = pw.isin
                                   AND srri.metric = 'srri_nav'
                                   AND srri.horizon = 'since_inception'
                                   AND srri.real_flag = 0
        WHERE pw.scenario_id = ?
        ORDER BY pw.block, pw.weight DESC
    """, (scenario_id,)).fetchall()

    sub_colors = {
        "Defensiva":   "D1ECF1",
        "Equilibrada": "FFF3CD",
        "Dinamica":    "F8D7DA",
    }

    GREEN_FILL = _fill("D4EDDA")
    RED_FILL   = _fill("F8D7DA")
    AMBER_FILL = _fill("FFF3CD")

    for r in rows:
        row_idx = ws.max_row + 1
        bg = sub_colors.get(r[0], "FFFFFF")
        bg_fill = _fill(bg)

        def _wc(col_n, val, fill=None):
            cell = ws.cell(row_idx, col_n, val)
            cell.font = _font(size=10)
            cell.fill = fill if fill else bg_fill
            if col_n in {8, 9, 10, 11, 12} and val is not None:
                try:
                    fv = float(val)
                    cell.value = fv / 100
                    cell.number_format = "0.00%"
                except Exception:
                    pass

        # Cols 1-9: datos base de la query
        for c_idx, val in enumerate(r[:9], 1):
            _wc(c_idx, val)

        # Col 10: Objetivo IPC+M3
        _wc(10, round(obj_anual, 2) if obj_anual else "N/D")

        # Col 11: Exceso vs objetivo
        ret_real = r[7]
        if ret_real is not None and obj_anual is not None:
            exceso = round(float(ret_real) - obj_anual, 2)
            _wc(11, exceso, GREEN_FILL if exceso >= 0 else RED_FILL)
        else:
            _wc(11, "N/D")

        # Col 12: Contribucion divisa anual %
        fx_val = r[9]
        _wc(12, fx_val if fx_val is not None else "",
            AMBER_FILL if (fx_val is not None and abs(float(fx_val)) > 20) else _fill(bg))

        # Cols 13-15: Moneda, Cobertura, Geografia
        _wc(13, r[10] or "")  # Fund_Currency
        _wc(14, r[11] or "")  # Hedging_Policy
        _wc(15, r[12] or "")  # Geography

        # Col 16: Beta Oil
        bo = r[13]
        _wc(16, bo, GREEN_FILL if (bo is not None and float(bo) > 0.01) else None)

        # Col 17: Beta Rate EU
        bre = r[14]
        _wc(17, bre, RED_FILL if (bre is not None and float(bre) < -0.10) else None)

        # Col 18: Beta Rate US
        _wc(18, r[15])

        # Col 19: Macro R2
        _wc(19, r[16])

        # Col 20: Alpha Persistence
        ap = r[17]
        _wc(20, ap,
            GREEN_FILL if (ap is not None and float(ap) >= 0.60)
            else AMBER_FILL if (ap is not None and float(ap) >= 0.40)
            else RED_FILL if ap is not None else None)

        # Col 21: SRRI KIID (r[18])
        srri_kiid = r[18]
        _wc(21, srri_kiid)

        # Col 22: SRRI Calculado (r[19])
        srri_calc = r[19]
        srri_bg = None
        if srri_kiid is not None and srri_calc is not None:
            try:
                diff = int(srri_calc) - int(srri_kiid)
                srri_bg = (RED_FILL if abs(diff) >= 2
                           else AMBER_FILL if abs(diff) == 1
                           else GREEN_FILL)
            except Exception:
                pass
        _wc(22, srri_calc, srri_bg)

    ws.auto_filter.ref = f"A2:{get_column_letter(22)}{ws.max_row}"
    ws.freeze_panes = "A3"
    _autofit(ws)


def _build_regimen(ws, clf):
    """Hoja 2 — Clasificacion macro actual e historica."""
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"A1:{get_column_letter(15)}1")
    ws["A1"] = "CLASIFICACION DE REGIMEN MACRO"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = _center()

    # Distribucion historica
    ws.cell(3, 1, "DISTRIBUCION HISTORICA 2000-2026").font = _font(bold=True, size=11)
    summary = clf.regime_summary()

    headers = ["Regimen", "Meses", "% Tiempo"]
    _apply_header(ws, headers, row=4)
    for i, (regime, row) in enumerate(summary.iterrows()):
        r_idx = 5 + i
        ws.cell(r_idx, 1, regime).font = _font(size=10)
        ws.cell(r_idx, 2, int(row["meses"])).font = _font(size=10)
        ws.cell(r_idx, 3, f"{row['pct']:.1f}%").font = _font(size=10)
        bg, fg = REGIME_COLORS.get(regime, ("FFFFFF", "000000"))
        for c in range(1, 4):
            ws.cell(r_idx, c).fill = _fill(bg)

    # Ultimos 24 meses
    hist = clf.classify_historical()
    if not hist.empty:
        ws.cell(ws.max_row + 2, 1,
                "CLASIFICACION MENSUAL (ultimos 120 meses)").font = _font(bold=True, size=11)
        headers2 = ["Fecha", "Regimen", "Defensiva", "Equilibrada", "Dinamica",
                    "IPC avg %", "CLI EU", "Tipo dep %", "Petroleo YoY %",
                    "M3 YoY %", "Tipo US %", "Tipo JP %", "Tipo CN %",
                    "Cobre YoY %", "D.Rate 3m"]
        _apply_header(ws, headers2, row=ws.max_row + 1)

        import numpy as _np
        macro_extra = clf._macro.ffill() if hasattr(clf, '_macro') else None

        def _v(row, col, mult=1, fmt=".1f"):
            """Obtiene valor de macro_df formateado, N/D si no disponible."""
            v = row.get(col) if hasattr(row, 'get') else None
            if v is None:
                return "N/D"
            try:
                fv = float(v)
                if _np.isnan(fv):
                    return "N/D"
                return format(fv * mult, fmt)
            except Exception:
                return "N/D"

        recent = hist.iloc[-120:]
        for dt, row in recent.iterrows():
            r_idx = ws.max_row + 1
            mr = macro_extra.loc[dt] if (macro_extra is not None
                                          and dt in macro_extra.index) else None

            def _mv(col, mult=1, fmt=".2f"):
                if mr is None:
                    return "N/D"
                v = mr.get(col)
                if v is None:
                    return "N/D"
                try:
                    fv = float(v)
                    return "N/D" if _np.isnan(fv) else format(fv * mult, fmt)
                except Exception:
                    return "N/D"

            vals = [
                dt.strftime("%Y-%m"),
                row["regime"],
                f"{row['weight_defensive']:.0%}",
                f"{row['weight_balanced']:.0%}",
                f"{row['weight_dynamic']:.0%}",
                _mv("ipc_yoy_avg", 100),
                _mv("cli_eu"),
                _mv("rate_deposit"),
                _mv("oil_yoy", 100, "+.1f"),
                _mv("m3_yoy"),
                _mv("rate_us"),
                _mv("rate_jp"),
                _mv("rate_cn"),
                _mv("copper_yoy", 1, "+.1f"),
                _mv("d_rate_3m"),
            ]
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(r_idx, c_idx, val)
                cell.font = _font(size=9)
            bg, _ = REGIME_COLORS.get(row["regime"], ("FFFFFF", "000000"))
            ws.cell(r_idx, 2).fill = _fill(bg)

    # Barras de color condicional para columnas numericas en regimen
    from openpyxl.formatting.rule import DataBarRule
    start_row = 13
    end_row   = ws.max_row
    if end_row >= start_row:
        for col_n in range(6, 16):
            col_letter = get_column_letter(col_n)
            cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
            rule = DataBarRule(start_type="min", start_value=0,
                               end_type="max", end_value=0,
                               color="638EC6")
            ws.conditional_formatting.add(cell_range, rule)

    # Nota sobre calidad de datos
    note_row = ws.max_row + 2
    ws.cell(note_row, 1,
            "(*) Tipo JP: serie IRSTCI01JPM156N (BoJ overnight). "
            "Tipo CN: INTDSRCNM193N (deposito), no refleja el LPR real "
            "(3.10% oct-2024, 3.00% may-2025). "
            "Datos con forward-fill cuando la fuente no tiene actualizacion mensual."
            ).font = _font(size=8, color="888780")
    ws.merge_cells(f"A{note_row}:{get_column_letter(15)}{note_row}")

    _autofit(ws)


def _build_backtesting(ws, conn):
    """Hoja 3 — Resumen del backtesting."""
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"A1:{get_column_letter(9)}1")
    ws["A1"] = "BACKTESTING -- Validacion historica del modelo (2005-2026)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = _center()

    bt = Backtester(conn)
    results = bt.run(start_date="2005-01-01")

    ws.cell(3, 1, "RESUMEN POR REGIMEN (horizonte 12 meses)").font = _font(bold=True, size=11)
    # Calcular objetivo IPC+M3 historico
    obj = _get_objetivo_ipc_m3(conn)
    obj_anual = obj["objetivo_anual"] if obj["objetivo_anual"] else 0.06

    # Añadir columna vs objetivo al results
    results["vs_objetivo"] = results["ret_12m"] - obj_anual

    headers = ["Regimen", "Meses", "Cartera 12m %",
               "Benchmark 12m %", "Exceso bench pp",
               "Objetivo IPC+M3 %", "Exceso objetivo pp",
               "Hit vs bench %", "Hit vs objetivo %"]
    _apply_header(ws, headers, row=4)

    for regime in results["regime"].unique():
        sub = results[results["regime"] == regime]
        valid = sub[["ret_12m", "bench_12m", "excess_12m", "vs_objetivo"]].dropna()
        if valid.empty:
            continue
        r_idx = ws.max_row + 1
        vals = [
            regime,
            len(sub),
            f"{valid['ret_12m'].mean()*100:+.1f}%",
            f"{valid['bench_12m'].mean()*100:+.1f}%",
            f"{valid['excess_12m'].mean()*100:+.1f}pp",
            f"{obj_anual*100:.1f}%",
            f"{valid['vs_objetivo'].mean()*100:+.1f}pp",
            f"{(valid['excess_12m'] > 0).mean()*100:.0f}%",
            f"{(valid['vs_objetivo'] > 0).mean()*100:.0f}%",
        ]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.font = _font(size=10)
        bg, _ = REGIME_COLORS.get(regime, ("FFFFFF", "000000"))
        ws.cell(r_idx, 1).fill = _fill(bg)

    # Global
    r_idx = ws.max_row + 2
    ws.cell(r_idx, 1, "GLOBAL").font = _font(bold=True, size=10)
    valid_all = results[["ret_12m", "bench_12m", "excess_12m"]].dropna()
    if not valid_all.empty:
        ws.cell(r_idx, 2, len(results))
        ws.cell(r_idx, 3, f"{valid_all['ret_12m'].mean()*100:+.1f}%").font = _font(bold=True, size=10)
        ws.cell(r_idx, 4, f"{valid_all['bench_12m'].mean()*100:+.1f}%").font = _font(bold=True, size=10)
        ws.cell(r_idx, 5, f"{valid_all['excess_12m'].mean()*100:+.1f}pp").font = _font(bold=True, size=10, color="155724")
        ws.cell(r_idx, 6, f"{obj_anual*100:.1f}%").font = _font(bold=True, size=10)
        vs_obj_all = valid_all["ret_12m"] - obj_anual
        ws.cell(r_idx, 7, f"{vs_obj_all.mean()*100:+.1f}pp").font = _font(bold=True, size=10,
            color="155724" if vs_obj_all.mean() > 0 else "721C24")
        ws.cell(r_idx, 8, f"{(valid_all['excess_12m'] > 0).mean()*100:.0f}%").font = _font(bold=True, size=10)
        ws.cell(r_idx, 9, f"{(vs_obj_all > 0).mean()*100:.0f}%").font = _font(bold=True, size=10)

    ws.cell(r_idx + 2, 1,
            "Nota: backtesting simple con look-ahead bias. "
            "Las metricas usan datos completos hasta la fecha de calculo.").font = _font(size=9, color="888780")

    _autofit(ws)


# ============================================================
# Generador principal
# ============================================================


def _build_macro_indicadores(ws, conn):
    """Hoja 4 — Indicadores macro: ultimos 36 meses con MM3 y MM6."""
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"A1:{get_column_letter(12)}1")
    ws["A1"] = "INDICADORES MACRO -- Evolucion 120 meses con medias moviles MM3 y MM6"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = _center()

    # Cargar datos macro de los ultimos 36 meses
    import pandas as _pd
    import numpy as _np

    indicadores = [
        ("ipc_index",    "ES",     "IPC Espana",        "YoY %",  True,  100),
        ("ipc_index",    "EU",     "IPC Eurozona",       "YoY %",  True,  100),
        ("m3_yoy",       "EU",     "M3 Eurozona",        "% YoY",  False, 1),
        ("rate_deposit", "EU",     "Tipo deposito BCE",  "%",      False, 1),
        ("rate_policy",  "US",     "Fed Funds",          "%",      False, 1),
        ("rate_policy",  "JP",     "Tipo BoJ",           "%",      False, 1),
        ("oil_wti",      "GLOBAL", "Petroleo WTI",       "YoY %",  True,  1),
        ("copper",       "GLOBAL", "Cobre",              "YoY %",  True,  1),
        ("cli",          "EU",     "CLI Europa (Alem.)", "Indice", False, 1),
        ("cli",          "US",     "CLI EEUU",           "Indice", False, 1),
        ("unemployment", "US",     "Desempleo EEUU",     "%",      False, 1),
    ]

    row_offset = 3

    for indicator, geography, label, unit, calc_yoy, mult in indicadores:
        rows_db = conn.execute("""
            SELECT date, value FROM series_macro
            WHERE indicator=? AND geography=?
            ORDER BY date DESC LIMIT 132
        """, (indicator, geography)).fetchall()

        if not rows_db:
            continue

        df = _pd.DataFrame(rows_db, columns=["date","value"])
        df["date"]  = _pd.to_datetime(df["date"])
        df["value"] = df["value"].astype(float)
        df = df.sort_values("date").tail(120)

        if calc_yoy and len(df) >= 13:
            df["plot"] = df["value"].pct_change(12) * 100
        else:
            df["plot"] = df["value"] * mult

        df = df.dropna(subset=["plot"]).tail(36)
        if df.empty:
            continue

        df["mm3"] = df["plot"].rolling(3).mean()
        df["mm6"] = df["plot"].rolling(6).mean()

        # Cabecera del indicador
        r_title = row_offset
        ws.cell(r_title, 1, f"{label} ({unit})").font = _font(bold=True, size=10)
        ws.cell(r_title, 1).fill = _fill("E6F1FB")

        headers = ["Fecha", "Valor", "MM3", "MM6", "Tendencia"]
        for c, h in enumerate(headers, 1):
            ws.cell(r_title+1, c, h).font = _font(bold=True, size=9)
            ws.cell(r_title+1, c).fill = _fill("457B9D")

        # Datos (ultimos 24 meses)
        recent = df.tail(120)
        for _, drow in recent.iterrows():
            r_idx = ws.max_row + 1 if ws.max_row >= r_title+1 else r_title+2
            r_idx = ws.max_row + 1

            val  = round(float(drow["plot"]), 2)
            mm3  = round(float(drow["mm3"]), 2) if not _np.isnan(drow["mm3"]) else None
            mm6  = round(float(drow["mm6"]), 2) if not _np.isnan(drow["mm6"]) else None

            # Semaforo tendencia
            if mm3 is not None and mm6 is not None:
                tendencia = "SUB" if mm3 > mm6 else "BAJ"
                trend_fill = _fill("D4EDDA") if mm3 > mm6 else _fill("F8D7DA")
            else:
                tendencia = "N/D"
                trend_fill = None

            ws.cell(r_idx, 1, drow["date"].strftime("%Y-%m")).font = _font(size=9)
            ws.cell(r_idx, 2, val).font = _font(size=9)
            ws.cell(r_idx, 3, mm3).font = _font(size=9)
            ws.cell(r_idx, 4, mm6).font = _font(size=9)
            tend_cell = ws.cell(r_idx, 5, tendencia)
            tend_cell.font = _font(size=9)
            if trend_fill:
                tend_cell.fill = trend_fill

        # Grafico desde columna G
        try:
            from openpyxl.chart import LineChart, Reference
            chart = LineChart()
            chart.title  = label
            chart.style    = 2
            chart.height   = 14   # relacion 3:4 visual (altura:anchura)
            chart.width    = 36   # hasta columna V aprox
            chart.grouping = "standard"
            # Eliminar borde exterior del area de trazado
            try:
                from openpyxl.drawing.fill import NoFill
                chart.plot_area.spPr = None
            except Exception:
                pass

            data_start = r_title + 2
            data_end   = ws.max_row
            if data_end > data_start + 3:
                for col_n, color, serie_title in [
                    (2, "1F77B4", "Valor"),
                    (3, "FF7F0E", "MM3"),
                    (4, "2CA02C", "MM6"),
                ]:
                    ref = Reference(ws, min_col=col_n, min_row=data_start,
                                    max_row=data_end)
                    chart.add_data(ref, titles_from_data=False)
                    idx_s = col_n - 2
                    # Inicializar SeriesLabel antes de asignar .v
                    if chart.series[idx_s].title is None:
                        from openpyxl.chart.series import SeriesLabel as _SL
                        chart.series[idx_s].title = _SL()
                    chart.series[idx_s].title.v = serie_title
                    chart.series[idx_s].graphicalProperties.line.solidFill = color
                    chart.series[idx_s].graphicalProperties.line.width = 18000

                dates_ref = Reference(ws, min_col=1, min_row=data_start,
                                      max_row=data_end)
                chart.set_categories(dates_ref)
                ws.add_chart(chart, f"G{r_title + 2}")
        except Exception:
            pass

        row_offset = ws.max_row + 2

    _autofit(ws)


def generate_report(
    conn: sqlite3.Connection,
    output_dir: str | None = None,
    scenario_id: str = "shock_energia_2026Q1",
    score_version: str = "v1",
) -> Path:
    """
    Genera el informe mensual completo en Excel.

    Devuelve la ruta del fichero generado.
    """
    today      = date.today().strftime("%Y%m%d")
    filename   = f"informe_cartera_{today}.xlsx"
    if output_dir is None:
        output_dir = str(_DEFAULT_REPORTS_DIR)
    output_path = Path(output_dir) / filename
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    clf = RegimeClassifier(conn)

    wb = Workbook()
    wb.remove(wb.active)  # eliminar hoja por defecto

    print("Generando informe mensual...")

    ws0 = wb.create_sheet("0_Portada")
    _build_portada(ws0, conn, clf, scenario_id)
    print("  0_Portada OK")

    ws1 = wb.create_sheet("1_Cartera")
    _build_cartera(ws1, conn, scenario_id)
    print("  1_Cartera OK")

    ws2 = wb.create_sheet("2_Regimen")
    _build_regimen(ws2, clf)
    print("  2_Regimen OK")

    ws3 = wb.create_sheet("3_Backtesting")
    _build_backtesting(ws3, conn)
    print("  3_Backtesting OK")

    ws4 = wb.create_sheet("4_Macro_Indicadores")
    _build_macro_indicadores(ws4, conn)
    print("  4_Macro_Indicadores OK")

    wb.save(output_path)
    print(f"\nInforme generado: {output_path}")
    return output_path
