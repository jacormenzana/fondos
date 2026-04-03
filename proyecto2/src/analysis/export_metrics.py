# proyecto2/src/analysis/export_metrics.py
# -*- coding: utf-8 -*-
"""
Exportacion de metricas P2 a Excel.

Genera un libro Excel con una hoja por bloque de analisis:
  0_Portada       Resumen del pipeline y parametros
  1_Estado        Estado general del pipeline
  2_SRRI          Distribucion y comparacion SRRI calculado vs KIID
  3_Rentabilidad  Distribucion y top fondos por rentabilidad real
  4_Riesgo        Distribucion drawdown y ratio retorno/drawdown
  5_Consistencia  Fondos mas consistentes
  6_Crisis        Comportamiento en periodos de crisis
  7_Candidatos    Pre-filtro de candidatos para P3

Uso:
    cd c:/desarrollo/fondos
    python -m proyecto2.src.analysis.export_metrics
    python -m proyecto2.src.analysis.export_metrics --output informes/
    python -m proyecto2.src.analysis.export_metrics --min-fondos 2500
"""

import argparse
import sqlite3
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle

_P2_SRC = Path(__file__).resolve().parent.parent.parent  # proyecto2/
_ROOT   = _P2_SRC.parent                                  # c:/desarrollo/fondos
sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(_P2_SRC))

from shared.config import DB_PATH, METRICS_DIR
from shared.db import get_connection

# ============================================================
# Estilos
# ============================================================

FONT_NAME = "Arial"
FONT_SIZE = 9

def _font(bold=False, color="000000", size=FONT_SIZE):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border_bottom():
    side = Side(style="thin", color="AAAAAA")
    return Border(bottom=side)

HEADER_FILL  = _fill("404040")
HEADER_FONT  = _font(bold=True, color="FFFFFF")
TITLE_FILL   = _fill("1F3864")
TITLE_FONT   = _font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = _fill("D6E4F0")
SECTION_FONT = _font(bold=True, color="1F3864")

# Colores condicionales
GREEN_FILL  = _fill("C6EFCE")
GREEN_FONT  = _font(color="276221")
RED_FILL    = _fill("FFC7CE")
RED_FONT    = _font(color="9C0006")
AMBER_FILL  = _fill("FFEB9C")
AMBER_FONT  = _font(color="9C5700")

# SRRI colores
SRRI_COLORS = {
    0: ("E2EFDA", "375623"),
    1: ("E2EFDA", "375623"),
    2: ("E2EFDA", "375623"),
    3: ("E2EFDA", "375623"),
    4: ("FFEB9C", "9C5700"),
    5: ("FFEB9C", "9C5700"),
    6: ("FFC7CE", "9C0006"),
    7: ("FFC7CE", "9C0006"),
}

# ============================================================
# Helpers de formato
# ============================================================

def _apply_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[row].height = 28

def _autofit(ws, min_width=8, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))

def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell

def _autofilter(ws):
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions

def _no_gridlines(ws):
    ws.sheet_view.showGridLines = False

def _write_rows(ws, rows, start_row=2, fmt_map=None):
    """
    Escribe filas de datos.
    fmt_map: dict {col_idx: format_string} para formato de celda
    """
    fmt_map = fmt_map or {}
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = _font()
            cell.alignment = Alignment(vertical="center")
            if c_idx in fmt_map:
                cell.number_format = fmt_map[c_idx]
    return r_idx if rows else start_row - 1

def _fmt_pct(val):
    """Convierte ratio a porcentaje redondeado con 2 decimales."""
    if val is None:
        return None
    return round(float(val) * 100, 2)

def _fmt_num(val, decimals=2):
    if val is None:
        return None
    return round(float(val), decimals)


# ============================================================
# Queries
# ============================================================

def q_estado(conn):
    return {
        "resumen": conn.execute("""
            SELECT COUNT(DISTINCT isin) AS fondos,
                   COUNT(*) AS metricas,
                   MIN(calculation_date) AS primera,
                   MAX(calculation_date) AS ultima
            FROM fund_metrics
        """).fetchall(),
        "sin_metricas": conn.execute("""
            SELECT COUNT(*) FROM fund_master fm
            WHERE NOT EXISTS (
                SELECT 1 FROM fund_metrics fmet WHERE fmet.isin = fm.ISIN
            )
        """).fetchone()[0],
        "por_horizonte": conn.execute("""
            SELECT horizon, COUNT(DISTINCT isin) AS fondos, COUNT(*) AS metricas
            FROM fund_metrics
            GROUP BY horizon ORDER BY fondos DESC
        """).fetchall(),
    }

def q_srri_distribucion(conn):
    return conn.execute("""
        SELECT CAST(fmet.value AS INTEGER) AS srri_calculado,
               COUNT(*) AS fondos
        FROM fund_metrics fmet
        WHERE fmet.metric='srri_nav' AND fmet.horizon='since_inception'
          AND fmet.real_flag=0
        GROUP BY CAST(fmet.value AS INTEGER)
        ORDER BY srri_calculado
    """).fetchall()

def q_srri_vs_kiid(conn):
    return conn.execute("""
        SELECT fmet.isin, fm.Fund_Name, fm.Fund_Nature, fm.Management_Company,
               fm.SRRI AS srri_kiid,
               CAST(fmet.value AS INTEGER) AS srri_calculado,
               ABS(fm.SRRI - fmet.value) AS diferencia
        FROM fund_metrics fmet
        JOIN fund_master fm ON fm.ISIN = fmet.isin
        WHERE fmet.metric='srri_nav' AND fmet.horizon='since_inception'
          AND fmet.real_flag=0 AND fm.SRRI IS NOT NULL
        ORDER BY diferencia DESC, fmet.isin
    """).fetchall()

def q_rentabilidad_dist(conn, real_flag):
    return conn.execute("""
        SELECT
            CASE
                WHEN value < -0.05 THEN '1. < -5%'
                WHEN value < 0     THEN '2. -5% a 0%'
                WHEN value < 0.03  THEN '3. 0% a 3%'
                WHEN value < 0.05  THEN '4. 3% a 5%'
                WHEN value < 0.08  THEN '5. 5% a 8%'
                WHEN value < 0.11  THEN '6. 8% a 11%'
                WHEN value < 0.15  THEN '7. 11% a 15%'
                ELSE                    '8. > 15%'
            END AS tramo,
            COUNT(*) AS fondos
        FROM fund_metrics
        WHERE metric='return_ann' AND horizon='since_inception'
          AND real_flag=? AND value IS NOT NULL
        GROUP BY tramo ORDER BY tramo
    """, (real_flag,)).fetchall()

def q_top_rentabilidad(conn):
    return conn.execute("""
        SELECT ret.isin, fm.Fund_Name, fm.Fund_Nature, fm.Management_Company,
               ROUND(ret.value*100,2)       AS return_real_pct,
               ROUND(vol.value*100,2)       AS volatilidad_pct,
               ROUND(sh.value,2)            AS sharpe,
               ROUND(srt.value,2)           AS sortino,
               ROUND(dd.value*100,2)        AS max_drawdown_pct,
               ROUND(pos.value*100,1)       AS pct_meses_positivos,
               CAST(srri.value AS INTEGER)  AS srri,
               ret.source_rows              AS meses_datos
        FROM fund_metrics ret
        JOIN fund_master fm   ON fm.ISIN=ret.isin
        LEFT JOIN fund_metrics vol  ON vol.isin=ret.isin  AND vol.metric='volatility_ann'
                                   AND vol.horizon='since_inception' AND vol.real_flag=0
        LEFT JOIN fund_metrics sh   ON sh.isin=ret.isin   AND sh.metric='sharpe'
                                   AND sh.horizon='since_inception' AND sh.real_flag=0
        LEFT JOIN fund_metrics srt  ON srt.isin=ret.isin  AND srt.metric='sortino'
                                   AND srt.horizon='since_inception' AND srt.real_flag=0
        LEFT JOIN fund_metrics dd   ON dd.isin=ret.isin   AND dd.metric='max_drawdown'
                                   AND dd.horizon='since_inception' AND dd.real_flag=0
        LEFT JOIN fund_metrics pos  ON pos.isin=ret.isin  AND pos.metric='pct_positive_months'
                                   AND pos.horizon='since_inception' AND pos.real_flag=0
        LEFT JOIN fund_metrics srri ON srri.isin=ret.isin AND srri.metric='srri_nav'
                                   AND srri.horizon='since_inception' AND srri.real_flag=0
        WHERE ret.metric='return_ann' AND ret.horizon='since_inception'
          AND ret.real_flag=1 AND ret.value IS NOT NULL
        ORDER BY ret.value DESC
    """).fetchall()

def q_drawdown_dist(conn):
    return conn.execute("""
        SELECT
            CASE
                WHEN value > -0.10 THEN '1. > -10%'
                WHEN value > -0.20 THEN '2. -10% a -20%'
                WHEN value > -0.30 THEN '3. -20% a -30%'
                WHEN value > -0.40 THEN '4. -30% a -40%'
                WHEN value > -0.50 THEN '5. -40% a -50%'
                ELSE                    '6. < -50%'
            END AS tramo,
            COUNT(*) AS fondos
        FROM fund_metrics
        WHERE metric='max_drawdown' AND horizon='since_inception'
          AND real_flag=0 AND value IS NOT NULL
        GROUP BY tramo ORDER BY tramo
    """).fetchall()

def q_ret_dd_ratio(conn):
    return conn.execute("""
        SELECT ret.isin, fm.Fund_Name, fm.Fund_Nature,
               ROUND(ret.value*100,2)              AS return_real_pct,
               ROUND(dd.value*100,2)               AS max_drawdown_pct,
               ROUND(ret.value/ABS(dd.value),2)    AS ret_dd_ratio,
               ROUND(sh.value,2)                   AS sharpe,
               CAST(srri.value AS INTEGER)          AS srri,
               ret.source_rows                     AS meses
        FROM fund_metrics ret
        JOIN fund_master fm   ON fm.ISIN=ret.isin
        JOIN fund_metrics dd  ON dd.isin=ret.isin  AND dd.metric='max_drawdown'
                             AND dd.horizon='since_inception' AND dd.real_flag=0
        LEFT JOIN fund_metrics sh   ON sh.isin=ret.isin  AND sh.metric='sharpe'
                                   AND sh.horizon='since_inception' AND sh.real_flag=0
        LEFT JOIN fund_metrics srri ON srri.isin=ret.isin AND srri.metric='srri_nav'
                                   AND srri.horizon='since_inception' AND srri.real_flag=0
        WHERE ret.metric='return_ann' AND ret.horizon='since_inception'
          AND ret.real_flag=1 AND ret.value>=0.02
          AND dd.value IS NOT NULL AND dd.value<0
        ORDER BY ret_dd_ratio DESC
    """).fetchall()

def q_consistencia(conn):
    return conn.execute("""
        SELECT pos.isin, fm.Fund_Name, fm.Fund_Nature,
               ROUND(pos.value*100,1)   AS pct_meses_positivos,
               ROUND(sev.value*100,1)   AS pct_perdida_severa,
               ROUND(wm.value*100,2)    AS peor_mes_pct,
               ROUND(ret.value*100,2)   AS return_real_pct,
               ROUND(sh.value,2)        AS sharpe,
               CAST(srri.value AS INTEGER) AS srri,
               pos.source_rows          AS meses
        FROM fund_metrics pos
        JOIN fund_master fm   ON fm.ISIN=pos.isin
        JOIN fund_metrics sev ON sev.isin=pos.isin AND sev.metric='pct_severe_loss_months'
                             AND sev.horizon='since_inception' AND sev.real_flag=0
        JOIN fund_metrics wm  ON wm.isin=pos.isin  AND wm.metric='worst_month'
                             AND wm.horizon='since_inception' AND wm.real_flag=0
        JOIN fund_metrics ret ON ret.isin=pos.isin  AND ret.metric='return_ann'
                             AND ret.horizon='since_inception' AND ret.real_flag=1
        LEFT JOIN fund_metrics sh   ON sh.isin=pos.isin  AND sh.metric='sharpe'
                                   AND sh.horizon='since_inception' AND sh.real_flag=0
        LEFT JOIN fund_metrics srri ON srri.isin=pos.isin AND srri.metric='srri_nav'
                                   AND srri.horizon='since_inception' AND srri.real_flag=0
        WHERE pos.metric='pct_positive_months'
          AND pos.horizon='since_inception' AND pos.real_flag=0
        ORDER BY pos.value DESC, sev.value ASC
    """).fetchall()

def q_crisis(conn):
    # Ancla en since_inception -- crisis_2020 es LEFT JOIN porque solo tiene 3 meses
    return conn.execute("""
        SELECT si.isin, fm.Fund_Name, fm.Fund_Nature,
               ROUND(c20.value*100,2)      AS return_crisis2020_pct,
               ROUND(c22.value*100,2)      AS return_crisis2022_pct,
               ROUND(c08.value*100,2)      AS return_crisis2008_pct,
               ROUND(c11.value*100,2)      AS return_crisis2011_pct,
               ROUND(si.value*100,2)       AS return_real_anual_pct,
               CAST(srri.value AS INTEGER) AS srri
        FROM fund_metrics si
        JOIN fund_master fm   ON fm.ISIN=si.isin
        LEFT JOIN fund_metrics c20 ON c20.isin=si.isin AND c20.metric='return_ann'
                                  AND c20.horizon='crisis_2020' AND c20.real_flag=0
        LEFT JOIN fund_metrics c22 ON c22.isin=si.isin AND c22.metric='return_ann'
                                  AND c22.horizon='crisis_2022' AND c22.real_flag=0
        LEFT JOIN fund_metrics c08 ON c08.isin=si.isin AND c08.metric='return_ann'
                                  AND c08.horizon='crisis_2008' AND c08.real_flag=0
        LEFT JOIN fund_metrics c11 ON c11.isin=si.isin AND c11.metric='return_ann'
                                  AND c11.horizon='crisis_2011' AND c11.real_flag=0
        LEFT JOIN fund_metrics srri ON srri.isin=si.isin AND srri.metric='srri_nav'
                                   AND srri.horizon='since_inception' AND srri.real_flag=0
        WHERE si.metric='return_ann' AND si.horizon='since_inception'
          AND si.real_flag=1 AND si.value IS NOT NULL
          AND (c22.value IS NOT NULL OR c08.value IS NOT NULL OR c11.value IS NOT NULL)
        ORDER BY (COALESCE(c22.value,0) + COALESCE(c08.value,0)) DESC
    """).fetchall()

def q_candidatos(conn):
    return conn.execute("""
        SELECT ret.isin, fm.Fund_Name, fm.Fund_Nature, fm.Management_Company,
               ROUND(ret.value*100,2)       AS return_real_pct,
               ROUND(vol.value*100,2)       AS volatilidad_pct,
               ROUND(sh.value,2)            AS sharpe,
               ROUND(srt.value,2)           AS sortino,
               ROUND(dd.value*100,2)        AS max_drawdown_pct,
               ROUND(pos.value*100,1)       AS pct_meses_positivos,
               ROUND(sev.value*100,1)       AS pct_perdida_severa,
               ROUND(wm.value*100,2)        AS peor_mes_pct,
               ROUND(oil.value, 4)          AS beta_oil,
               ROUND(cop.value, 4)          AS beta_copper,
               CAST(srri.value AS INTEGER)  AS srri_calculado,
               fm.SRRI                      AS srri_kiid,
               ret.source_rows              AS meses_datos
        FROM fund_metrics ret
        JOIN fund_master fm   ON fm.ISIN=ret.isin
        JOIN fund_metrics vol ON vol.isin=ret.isin  AND vol.metric='volatility_ann'
                             AND vol.horizon='since_inception' AND vol.real_flag=0
        JOIN fund_metrics sh  ON sh.isin=ret.isin   AND sh.metric='sharpe'
                             AND sh.horizon='since_inception' AND sh.real_flag=0
        JOIN fund_metrics srt ON srt.isin=ret.isin  AND srt.metric='sortino'
                             AND srt.horizon='since_inception' AND srt.real_flag=0
        JOIN fund_metrics dd  ON dd.isin=ret.isin   AND dd.metric='max_drawdown'
                             AND dd.horizon='since_inception' AND dd.real_flag=0
        JOIN fund_metrics pos ON pos.isin=ret.isin  AND pos.metric='pct_positive_months'
                             AND pos.horizon='since_inception' AND pos.real_flag=0
        JOIN fund_metrics sev ON sev.isin=ret.isin  AND sev.metric='pct_severe_loss_months'
                             AND sev.horizon='since_inception' AND sev.real_flag=0
        JOIN fund_metrics wm  ON wm.isin=ret.isin   AND wm.metric='worst_month'
                             AND wm.horizon='since_inception' AND wm.real_flag=0
        LEFT JOIN fund_metrics oil  ON oil.isin=ret.isin  AND oil.metric='beta_oil'
                                   AND oil.horizon='since_inception' AND oil.real_flag=0
        LEFT JOIN fund_metrics cop  ON cop.isin=ret.isin  AND cop.metric='beta_copper'
                                   AND cop.horizon='since_inception' AND cop.real_flag=0
        LEFT JOIN fund_metrics srri ON srri.isin=ret.isin AND srri.metric='srri_nav'
                                   AND srri.horizon='since_inception' AND srri.real_flag=0
        WHERE ret.metric='return_ann' AND ret.horizon='since_inception'
          AND ret.real_flag=1
          AND ret.value  >= 0.05
          AND sh.value   >= 0
          AND dd.value   >= -0.40
          AND pos.value  >= 0.55
          AND ret.source_rows >= 36
        ORDER BY ret.value DESC
    """).fetchall()


# ============================================================
# Constructores de hojas
# ============================================================

def build_portada(ws, conn, ts_str):
    _no_gridlines(ws)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 45

    fondos = conn.execute(
        "SELECT COUNT(DISTINCT isin) FROM fund_metrics").fetchone()[0]
    metricas = conn.execute(
        "SELECT COUNT(*) FROM fund_metrics").fetchone()[0]
    fondos_total = conn.execute(
        "SELECT COUNT(*) FROM fund_master").fetchone()[0]

    datos = [
        ("INFORME DE METRICAS P2", "", True),
        ("", "", False),
        ("Fecha de generacion",  ts_str, False),
        ("Fondos en universo",   fondos_total, False),
        ("Fondos con metricas",  fondos, False),
        ("Cobertura",            f"{fondos/fondos_total*100:.1f}%", False),
        ("Total metricas",       f"{metricas:,}", False),
        ("Metricas por fondo",   f"~{metricas//fondos if fondos else 0}", False),
        ("", "", False),
        ("HOJAS DEL INFORME", "", True),
        ("1_Estado",        "Estado general del pipeline", False),
        ("2_SRRI",          "Distribucion SRRI calculado vs KIID", False),
        ("3_Rentabilidad",  "Top fondos por rentabilidad real anualizada", False),
        ("4_Riesgo",        "Distribucion drawdown y ratio retorno/drawdown", False),
        ("5_Consistencia",  "Fondos mas consistentes", False),
        ("6_Crisis",        "Comportamiento en periodos de crisis", False),
        ("7_Candidatos",    "Pre-filtro de candidatos para P3", False),
        ("8_Macro_Betas",   "Sensibilidades macro individuales (betas OLS)", False),
    ]

    for r, (col_a, col_b, is_title) in enumerate(datos, 1):
        ca = ws.cell(row=r, column=1, value=col_a)
        cb = ws.cell(row=r, column=2, value=col_b)
        if is_title:
            for c in (ca, cb):
                c.font  = TITLE_FONT
                c.fill  = TITLE_FILL
            ws.row_dimensions[r].height = 22
        else:
            ca.font = _font(bold=True)
            cb.font = _font()


def build_estado(ws, conn):
    _no_gridlines(ws)

    # Resumen
    ws.merge_cells("A1:D1")
    ws["A1"] = "ESTADO GENERAL DEL PIPELINE"
    ws["A1"].font  = TITLE_FONT
    ws["A1"].fill  = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    _apply_header(ws, ["Fondos con metricas", "Total metricas",
                       "Primera calculo", "Ultimo calculo"], row=2)
    data = q_estado(conn)
    for r in data["resumen"]:
        ws.append(list(r))
        for c in range(1, 5):
            ws.cell(ws.max_row, c).font = _font()

    ws.append([])
    row_sep = ws.max_row + 1
    ws.merge_cells(f"A{row_sep}:B{row_sep}")
    ws.cell(row_sep, 1).value = f"Fondos sin metricas: {data['sin_metricas']}"
    ws.cell(row_sep, 1).font  = _font(bold=True)

    ws.append([])
    _apply_header(ws, ["Horizonte", "Fondos", "Metricas"], row=ws.max_row + 1)
    for r in data["por_horizonte"]:
        ws.append(list(r))
        for c in range(1, 4):
            ws.cell(ws.max_row, c).font = _font()

    _autofit(ws)
    _freeze(ws, "A3")


def build_srri(ws, conn):
    _no_gridlines(ws)

    ws.merge_cells("A1:C1")
    ws["A1"] = "DISTRIBUCION SRRI CALCULADO"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    _apply_header(ws, ["SRRI calculado", "Fondos"], row=2)
    dist = q_srri_distribucion(conn)
    for srri_val, fondos in dist:
        row_idx = ws.max_row + 1
        ws.cell(row_idx, 1, srri_val).font = _font(bold=True)
        ws.cell(row_idx, 2, fondos).font   = _font()
        bg, fg = SRRI_COLORS.get(srri_val or 0, ("FFFFFF", "000000"))
        ws.cell(row_idx, 1).fill = _fill(bg)
        ws.cell(row_idx, 1).font = _font(bold=True, color=fg)

    ws.append([])
    ws.append([])
    ws.merge_cells(f"A{ws.max_row}:G{ws.max_row}")
    ws.cell(ws.max_row, 1).value = "COMPARACION SRRI CALCULADO vs KIID"
    ws.cell(ws.max_row, 1).font  = TITLE_FONT
    ws.cell(ws.max_row, 1).fill  = TITLE_FILL
    ws.cell(ws.max_row, 1).alignment = Alignment(horizontal="center")

    headers = ["ISIN", "Nombre fondo", "Naturaleza", "Gestora",
               "SRRI KIID", "SRRI calculado", "Diferencia"]
    _apply_header(ws, headers, row=ws.max_row + 1)
    header_row = ws.max_row

    rows = q_srri_vs_kiid(conn)
    for r in rows:
        row_idx = ws.max_row + 1
        for c_idx, val in enumerate(r, 1):
            cell = ws.cell(row_idx, c_idx, val)
            cell.font = _font()
        # Color segun diferencia
        diff = r[6]
        if diff is not None:
            if diff == 0:
                ws.cell(row_idx, 7).fill = GREEN_FILL
                ws.cell(row_idx, 7).font = GREEN_FONT
            elif diff == 1:
                ws.cell(row_idx, 7).fill = AMBER_FILL
                ws.cell(row_idx, 7).font = AMBER_FONT
            else:
                ws.cell(row_idx, 7).fill = RED_FILL
                ws.cell(row_idx, 7).font = RED_FONT

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(7)}{ws.max_row}"
    _freeze(ws, f"A{header_row+1}")
    _autofit(ws)


def build_rentabilidad(ws, conn):
    _no_gridlines(ws)

    # Distribuciones
    ws.merge_cells("A1:B1")
    ws["A1"] = "Rentabilidad nominal since_inception"
    ws["A1"].font = SECTION_FONT
    ws["A1"].fill = SECTION_FILL
    _apply_header(ws, ["Tramo", "Fondos"], row=2)
    for r in q_rentabilidad_dist(conn, 0):
        ws.append(list(r))
        for c in range(1, 3):
            ws.cell(ws.max_row, c).font = _font()

    ws.cell(1, 4).value = "Rentabilidad REAL since_inception"
    ws.cell(1, 4).font  = SECTION_FONT
    ws.cell(1, 4).fill  = SECTION_FILL
    ws.cell(2, 4).value = "Tramo"
    ws.cell(2, 4).font  = HEADER_FONT
    ws.cell(2, 4).fill  = HEADER_FILL
    ws.cell(2, 5).value = "Fondos"
    ws.cell(2, 5).font  = HEADER_FONT
    ws.cell(2, 5).fill  = HEADER_FILL
    for i, r in enumerate(q_rentabilidad_dist(conn, 1), 3):
        ws.cell(i, 4, r[0]).font = _font()
        ws.cell(i, 5, r[1]).font = _font()

    # Top fondos
    sep_row = 14
    ws.merge_cells(f"A{sep_row}:{get_column_letter(12)}{sep_row}")
    ws.cell(sep_row, 1).value = "TOP FONDOS POR RENTABILIDAD REAL ANUALIZADA"
    ws.cell(sep_row, 1).font  = TITLE_FONT
    ws.cell(sep_row, 1).fill  = TITLE_FILL
    ws.cell(sep_row, 1).alignment = Alignment(horizontal="center")

    headers = ["ISIN", "Nombre", "Naturaleza", "Gestora",
               "Rent. real %", "Volatilidad %", "Sharpe", "Sortino",
               "Max DD %", "% Meses+", "SRRI", "Meses datos"]
    _apply_header(ws, headers, row=sep_row + 1)
    header_row = sep_row + 1

    fmt_map = {5: "0.00", 6: "0.00", 7: "0.00", 8: "0.00",
               9: "0.00", 10: "0.00"}

    rows = q_top_rentabilidad(conn)
    for r in rows:
        row_idx = ws.max_row + 1
        for c_idx, val in enumerate(r, 1):
            cell = ws.cell(row_idx, c_idx, val)
            cell.font = _font()
            if c_idx in fmt_map:
                cell.number_format = fmt_map[c_idx]
        # Color rentabilidad
        ret_cell = ws.cell(row_idx, 5)
        if ret_cell.value is not None:
            if float(ret_cell.value) >= 11:
                ret_cell.fill = GREEN_FILL
                ret_cell.font = GREEN_FONT
            elif float(ret_cell.value) < 0:
                ret_cell.fill = RED_FILL
                ret_cell.font = RED_FONT
        # Color SRRI
        srri_val = r[10]
        if srri_val is not None:
            bg, fg = SRRI_COLORS.get(int(srri_val), ("FFFFFF", "000000"))
            ws.cell(row_idx, 11).fill = _fill(bg)
            ws.cell(row_idx, 11).font = _font(color=fg)

    ws.auto_filter.ref = (f"A{header_row}:{get_column_letter(12)}{ws.max_row}")
    _freeze(ws, f"A{header_row+1}")
    _autofit(ws)


def build_riesgo(ws, conn):
    _no_gridlines(ws)

    ws.merge_cells("A1:B1")
    ws["A1"] = "Distribucion Max Drawdown nominal"
    ws["A1"].font = SECTION_FONT
    ws["A1"].fill = SECTION_FILL
    _apply_header(ws, ["Tramo", "Fondos"], row=2)
    for r in q_drawdown_dist(conn):
        ws.append(list(r))
        for c in range(1, 3):
            ws.cell(ws.max_row, c).font = _font()

    sep_row = 12
    ws.merge_cells(f"A{sep_row}:{get_column_letter(9)}{sep_row}")
    ws.cell(sep_row, 1).value = "RANKING POR RATIO RETORNO / MAX DRAWDOWN"
    ws.cell(sep_row, 1).font  = TITLE_FONT
    ws.cell(sep_row, 1).fill  = TITLE_FILL
    ws.cell(sep_row, 1).alignment = Alignment(horizontal="center")

    headers = ["ISIN", "Nombre", "Naturaleza",
               "Rent. real %", "Max DD %", "Ratio Ret/DD",
               "Sharpe", "SRRI", "Meses"]
    _apply_header(ws, headers, row=sep_row + 1)
    header_row = sep_row + 1

    rows = q_ret_dd_ratio(conn)
    for r in rows:
        row_idx = ws.max_row + 1
        for c_idx, val in enumerate(r, 1):
            cell = ws.cell(row_idx, c_idx, val)
            cell.font = _font()
        srri_val = r[7]
        if srri_val is not None:
            bg, fg = SRRI_COLORS.get(int(srri_val), ("FFFFFF", "000000"))
            ws.cell(row_idx, 8).fill = _fill(bg)
            ws.cell(row_idx, 8).font = _font(color=fg)

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(9)}{ws.max_row}"
    _freeze(ws, f"A{header_row+1}")
    _autofit(ws)


def build_consistencia(ws, conn):
    _no_gridlines(ws)

    ws.merge_cells(f"A1:{get_column_letter(10)}1")
    ws["A1"] = "CONSISTENCIA — FONDOS MAS ESTABLES"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["ISIN", "Nombre", "Naturaleza",
               "% Meses+", "% Perdida severa", "Peor mes %",
               "Rent. real %", "Sharpe", "SRRI", "Meses"]
    _apply_header(ws, headers, row=2)

    rows = q_consistencia(conn)
    for r in rows:
        row_idx = ws.max_row + 1
        for c_idx, val in enumerate(r, 1):
            cell = ws.cell(row_idx, c_idx, val)
            cell.font = _font()
        srri_val = r[8]
        if srri_val is not None:
            bg, fg = SRRI_COLORS.get(int(srri_val), ("FFFFFF", "000000"))
            ws.cell(row_idx, 9).fill = _fill(bg)
            ws.cell(row_idx, 9).font = _font(color=fg)

    ws.auto_filter.ref = f"A2:{get_column_letter(10)}{ws.max_row}"
    _freeze(ws, "A3")
    _autofit(ws)


def build_crisis(ws, conn):
    _no_gridlines(ws)

    ws.merge_cells(f"A1:{get_column_letter(9)}1")
    ws["A1"] = "COMPORTAMIENTO EN PERIODOS DE CRISIS"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["ISIN", "Nombre", "Naturaleza",
               "Crisis 2020 %", "Crisis 2022 %",
               "Crisis 2008 %", "Crisis 2011 %",
               "Rent. real anual %", "SRRI"]
    _apply_header(ws, headers, row=2)

    rows = q_crisis(conn)
    for r in rows:
        row_idx = ws.max_row + 1
        for c_idx, val in enumerate(r, 1):
            cell = ws.cell(row_idx, c_idx, val)
            cell.font = _font()
            if c_idx in (4, 5, 6, 7) and val is not None:
                if float(val) >= 0:
                    cell.fill = GREEN_FILL
                    cell.font = GREEN_FONT
                elif float(val) < -10:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT
        srri_val = r[8]
        if srri_val is not None:
            bg, fg = SRRI_COLORS.get(int(srri_val), ("FFFFFF", "000000"))
            ws.cell(row_idx, 9).fill = _fill(bg)
            ws.cell(row_idx, 9).font = _font(color=fg)

    ws.auto_filter.ref = f"A2:{get_column_letter(9)}{ws.max_row}"
    _freeze(ws, "A3")
    _autofit(ws)



def q_macro_betas(conn):
    return conn.execute("""
        SELECT ret.isin, fm.Fund_Name, fm.Fund_Nature,
               ROUND(r2.value, 3)           AS macro_r2,
               ROUND(alp.value*100, 2)      AS macro_alpha_pct,
               ROUND(reu.value, 4)          AS beta_rate_eu,
               ROUND(rus.value, 4)          AS beta_rate_us,
               ROUND(rjp.value, 4)          AS beta_rate_jp,
               ROUND(rcn.value, 4)          AS beta_rate_cn,
               ROUND(m3.value, 4)           AS beta_m3,
               ROUND(ies.value, 4)          AS beta_ipc_es,
               ROUND(ieu.value, 4)          AS beta_ipc_eu,
               ROUND(ius.value, 4)          AS beta_ipc_us,
               ROUND(ijp.value, 4)          AS beta_ipc_jp,
               ROUND(icn.value, 4)          AS beta_ipc_cn,
               ROUND(oil.value, 4)          AS beta_oil,
               ROUND(cop.value, 4)          AS beta_copper,
               ROUND(clieu.value, 4)        AS beta_cli_eu,
               ROUND(clus.value, 4)         AS beta_cli_us,
               ROUND(dxy.value, 4)          AS beta_dxy,
               ROUND(gld.value, 4)          AS beta_gold,
               ROUND(m2g.value, 4)          AS beta_m2_global,
               ROUND(sph.value, 4)          AS beta_spread_hy,
               ROUND(vix.value, 4)          AS beta_vix,
               ROUND(tsp.value, 4)          AS beta_term_spread,
               ROUND(ejy.value, 4)          AS beta_eur_jpy,
               ROUND(egb.value, 4)          AS beta_eur_gbp,
               ROUND(ecn.value, 4)          AS beta_eur_cny,
               CAST(srri.value AS INTEGER)  AS srri,
               CAST(nobs.value AS INTEGER)  AS n_obs
        FROM fund_metrics ret
        JOIN fund_master fm    ON fm.ISIN=ret.isin
        LEFT JOIN fund_metrics r2   ON r2.isin=ret.isin   AND r2.metric='macro_r2'
                                   AND r2.horizon='since_inception' AND r2.real_flag=0
        LEFT JOIN fund_metrics alp  ON alp.isin=ret.isin  AND alp.metric='macro_alpha'
                                   AND alp.horizon='since_inception' AND alp.real_flag=0
        LEFT JOIN fund_metrics reu  ON reu.isin=ret.isin  AND reu.metric='beta_rate_eu'
                                   AND reu.horizon='since_inception' AND reu.real_flag=0
        LEFT JOIN fund_metrics rus  ON rus.isin=ret.isin  AND rus.metric='beta_rate_us'
                                   AND rus.horizon='since_inception' AND rus.real_flag=0
        LEFT JOIN fund_metrics rjp  ON rjp.isin=ret.isin  AND rjp.metric='beta_rate_jp'
                                   AND rjp.horizon='since_inception' AND rjp.real_flag=0
        LEFT JOIN fund_metrics rcn  ON rcn.isin=ret.isin  AND rcn.metric='beta_rate_cn'
                                   AND rcn.horizon='since_inception' AND rcn.real_flag=0
        LEFT JOIN fund_metrics m3   ON m3.isin=ret.isin   AND m3.metric='beta_m3_yoy'
                                   AND m3.horizon='since_inception' AND m3.real_flag=0
        LEFT JOIN fund_metrics ies  ON ies.isin=ret.isin  AND ies.metric='beta_ipc_es'
                                   AND ies.horizon='since_inception' AND ies.real_flag=0
        LEFT JOIN fund_metrics ieu  ON ieu.isin=ret.isin  AND ieu.metric='beta_ipc_eu'
                                   AND ieu.horizon='since_inception' AND ieu.real_flag=0
        LEFT JOIN fund_metrics ius  ON ius.isin=ret.isin  AND ius.metric='beta_ipc_us'
                                   AND ius.horizon='since_inception' AND ius.real_flag=0
        LEFT JOIN fund_metrics ijp  ON ijp.isin=ret.isin  AND ijp.metric='beta_ipc_jp'
                                   AND ijp.horizon='since_inception' AND ijp.real_flag=0
        LEFT JOIN fund_metrics icn  ON icn.isin=ret.isin  AND icn.metric='beta_ipc_cn'
                                   AND icn.horizon='since_inception' AND icn.real_flag=0
        LEFT JOIN fund_metrics oil  ON oil.isin=ret.isin  AND oil.metric='beta_oil'
                                   AND oil.horizon='since_inception' AND oil.real_flag=0
        LEFT JOIN fund_metrics cop  ON cop.isin=ret.isin  AND cop.metric='beta_copper'
                                   AND cop.horizon='since_inception' AND cop.real_flag=0
        LEFT JOIN fund_metrics clieu ON clieu.isin=ret.isin AND clieu.metric='beta_cli_eu'
                                   AND clieu.horizon='since_inception' AND clieu.real_flag=0
        LEFT JOIN fund_metrics clus  ON clus.isin=ret.isin  AND clus.metric='beta_cli_us'
                                   AND clus.horizon='since_inception' AND clus.real_flag=0
        LEFT JOIN fund_metrics dxy  ON dxy.isin=ret.isin   AND dxy.metric='beta_dxy'
                                   AND dxy.horizon='since_inception' AND dxy.real_flag=0
        LEFT JOIN fund_metrics gld  ON gld.isin=ret.isin   AND gld.metric='beta_gold'
                                   AND gld.horizon='since_inception' AND gld.real_flag=0
        LEFT JOIN fund_metrics m2g  ON m2g.isin=ret.isin   AND m2g.metric='beta_m2_global'
                                   AND m2g.horizon='since_inception' AND m2g.real_flag=0
        LEFT JOIN fund_metrics sph  ON sph.isin=ret.isin   AND sph.metric='beta_spread_hy'
                                   AND sph.horizon='since_inception' AND sph.real_flag=0
        LEFT JOIN fund_metrics vix  ON vix.isin=ret.isin   AND vix.metric='beta_vix'
                                   AND vix.horizon='since_inception' AND vix.real_flag=0
        LEFT JOIN fund_metrics tsp  ON tsp.isin=ret.isin   AND tsp.metric='beta_term_spread'
                                   AND tsp.horizon='since_inception' AND tsp.real_flag=0
        LEFT JOIN fund_metrics ejy  ON ejy.isin=ret.isin   AND ejy.metric='beta_eur_jpy'
                                   AND ejy.horizon='since_inception' AND ejy.real_flag=0
        LEFT JOIN fund_metrics egb  ON egb.isin=ret.isin   AND egb.metric='beta_eur_gbp'
                                   AND egb.horizon='since_inception' AND egb.real_flag=0
        LEFT JOIN fund_metrics ecn  ON ecn.isin=ret.isin   AND ecn.metric='beta_eur_cny'
                                   AND ecn.horizon='since_inception' AND ecn.real_flag=0
        LEFT JOIN fund_metrics srri ON srri.isin=ret.isin AND srri.metric='srri_nav'
                                   AND srri.horizon='since_inception' AND srri.real_flag=0
        LEFT JOIN fund_metrics nobs ON nobs.isin=ret.isin AND nobs.metric='macro_n_obs'
                                   AND nobs.horizon='since_inception' AND nobs.real_flag=0
        WHERE ret.metric='return_ann' AND ret.horizon='since_inception'
          AND ret.real_flag=0 AND r2.value IS NOT NULL
        ORDER BY r2.value DESC
    """).fetchall()


def build_macro(ws, conn):
    _no_gridlines(ws)

    N_COLS = 30
    ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
    ws["A1"] = "SENSIBILIDADES MACRO POR FONDO (betas OLS) — pipeline v10: 23 factores"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "ISIN", "Nombre", "Naturaleza",
        "R2 macro", "Alpha anual %",
        # Tipos
        "β rate EU", "β rate US", "β rate JP", "β rate CN",
        # Monetario
        "β M3", "β M2 Global",
        # Inflación
        "β IPC ES", "β IPC EU", "β IPC US", "β IPC JP", "β IPC CN",
        # Materias primas
        "β Oil", "β Cobre", "β Oro",
        # Ciclo / FX
        "β CLI EU", "β CLI US", "β DXY",
        # Riesgo financiero (v10)
        "β Spread HY", "β VIX", "β Term Spread",
        # Divisas (v10)
        "β EUR/JPY", "β EUR/GBP", "β EUR/CNY",
        # Control
        "SRRI", "N obs",
    ]
    _apply_header(ws, headers, row=2)

    rows = q_macro_betas(conn)
    for r in rows:
        row_idx = ws.max_row + 1
        for c_idx, val in enumerate(r, 1):
            cell = ws.cell(row_idx, c_idx, val)
            cell.font = _font()
        # Color R2 (columna 4)
        r2_val = r[3]
        if r2_val is not None:
            cell = ws.cell(row_idx, 4)
            if float(r2_val) >= 0.50:
                cell.fill = RED_FILL
                cell.font = RED_FONT
            elif float(r2_val) >= 0.30:
                cell.fill = AMBER_FILL
                cell.font = AMBER_FONT
            elif float(r2_val) >= 0.10:
                cell.fill = GREEN_FILL
                cell.font = GREEN_FONT
        # SRRI (columna 29)
        srri_val = r[28]
        if srri_val is not None:
            bg, fg = SRRI_COLORS.get(int(srri_val), ("FFFFFF", "000000"))
            ws.cell(row_idx, 29).fill = _fill(bg)
            ws.cell(row_idx, 29).font = _font(color=fg)

    ws.auto_filter.ref = f"A2:{get_column_letter(N_COLS)}{ws.max_row}"
    _freeze(ws, "A3")
    _autofit(ws)

    ws.auto_filter.ref = f"A2:{get_column_letter(21)}{ws.max_row}"
    _freeze(ws, "A3")
    _autofit(ws)

def build_candidatos(ws, conn):
    _no_gridlines(ws)

    ws.merge_cells(f"A1:{get_column_letter(15)}1")
    ws["A1"] = "CANDIDATOS P3 — Retorno real>=5%, Sharpe>=0, MaxDD>=-40%, Meses+>=55%, Min 36 meses"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["ISIN", "Nombre", "Naturaleza", "Gestora",
               "Rent. real %", "Volatilidad %", "Sharpe", "Sortino",
               "Max DD %", "% Meses+", "% Perdida sev.",
               "Peor mes %", "Beta Oil", "Beta Cobre",
               "SRRI calc.", "SRRI KIID", "Meses datos"]
    _apply_header(ws, headers, row=2)

    rows = q_candidatos(conn)
    for r in rows:
        row_idx = ws.max_row + 1
        for c_idx, val in enumerate(r, 1):
            cell = ws.cell(row_idx, c_idx, val)
            cell.font = _font()
        ret_val = r[4]
        if ret_val is not None:
            cell = ws.cell(row_idx, 5)
            if float(ret_val) >= 11:
                cell.fill = GREEN_FILL
                cell.font = GREEN_FONT
            elif float(ret_val) >= 5:
                cell.fill = AMBER_FILL
                cell.font = AMBER_FONT
        srri_val = r[12]
        if srri_val is not None:
            bg, fg = SRRI_COLORS.get(int(srri_val), ("FFFFFF", "000000"))
            ws.cell(row_idx, 13).fill = _fill(bg)
            ws.cell(row_idx, 13).font = _font(color=fg)

    # Conteo por naturaleza al final
    ws.append([])
    summary_row = ws.max_row + 1
    ws.cell(summary_row, 1).value = "Candidatos por naturaleza de fondo:"
    ws.cell(summary_row, 1).font  = _font(bold=True)

    by_nature = {}
    for r in rows:
        nat = r[2] or "Sin clasificar"
        by_nature[nat] = by_nature.get(nat, 0) + 1
    for nat, cnt in sorted(by_nature.items(), key=lambda x: -x[1]):
        ws.append([nat, cnt])
        ws.cell(ws.max_row, 1).font = _font()
        ws.cell(ws.max_row, 2).font = _font()

    ws.auto_filter.ref = f"A2:{get_column_letter(15)}{ws.max_row}"
    _freeze(ws, "A3")
    _autofit(ws)


# ============================================================
# Orquestador principal
# ============================================================


def q_persistencia(conn):
    return conn.execute("""
        SELECT per.isin, fm.Fund_Name, fm.Fund_Nature, fm.Management_Company,
               ROUND(per.value, 3)              AS alpha_persistence,
               CAST(nobs.value AS INTEGER)       AS n_ventanas,
               ROUND(ret.value*100, 2)           AS return_real_pct,
               ROUND(sh.value, 2)                AS sharpe,
               ROUND(dd.value*100, 2)            AS max_dd_pct,
               CAST(srri.value AS INTEGER)        AS srri
        FROM fund_metrics per
        JOIN fund_master fm    ON fm.ISIN=per.isin
        LEFT JOIN fund_metrics nobs ON nobs.isin=per.isin AND nobs.metric='alpha_persistence_n'
                                   AND nobs.horizon='since_inception' AND nobs.real_flag=0
        LEFT JOIN fund_metrics ret  ON ret.isin=per.isin  AND ret.metric='return_ann'
                                   AND ret.horizon='since_inception' AND ret.real_flag=1
        LEFT JOIN fund_metrics sh   ON sh.isin=per.isin   AND sh.metric='sharpe'
                                   AND sh.horizon='since_inception' AND sh.real_flag=0
        LEFT JOIN fund_metrics dd   ON dd.isin=per.isin   AND dd.metric='max_drawdown'
                                   AND dd.horizon='since_inception' AND dd.real_flag=0
        LEFT JOIN fund_metrics srri ON srri.isin=per.isin AND srri.metric='srri_nav'
                                   AND srri.horizon='since_inception' AND srri.real_flag=0
        WHERE per.metric='alpha_persistence' AND per.horizon='since_inception'
          AND per.real_flag=0 AND per.value IS NOT NULL
        ORDER BY per.value DESC
    """).fetchall()


def build_persistencia(ws, conn):
    _no_gridlines(ws)
    ws.merge_cells(f"A1:{get_column_letter(10)}1")
    ws["A1"] = "PERSISTENCIA DEL ALPHA -- % ventanas rolling 3Y superando a la categoria"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["ISIN", "Nombre", "Naturaleza", "Gestora",
               "Persistencia", "N ventanas", "Rent. real %",
               "Sharpe", "Max DD %", "SRRI"]
    _apply_header(ws, headers, row=2)

    rows = q_persistencia(conn)
    for r in rows:
        row_idx = ws.max_row + 1
        for c_idx, val in enumerate(r, 1):
            cell = ws.cell(row_idx, c_idx, val)
            cell.font = _font()
        pers_val = r[4]
        if pers_val is not None:
            cell = ws.cell(row_idx, 5)
            if float(pers_val) >= 0.60:
                cell.fill = GREEN_FILL; cell.font = GREEN_FONT
            elif float(pers_val) >= 0.40:
                cell.fill = AMBER_FILL; cell.font = AMBER_FONT
            else:
                cell.fill = RED_FILL;  cell.font = RED_FONT
        srri_val = r[9]
        if srri_val is not None:
            bg, fg = SRRI_COLORS.get(int(srri_val), ("FFFFFF","000000"))
            ws.cell(row_idx, 10).fill = _fill(bg)
            ws.cell(row_idx, 10).font = _font(color=fg)

    ws.auto_filter.ref = f"A2:{get_column_letter(10)}{ws.max_row}"
    _freeze(ws, "A3")
    _autofit(ws)


def q_divisa(conn):
    return conn.execute("""
        SELECT fx.isin, fm.Fund_Name, fm.Fund_Nature,
               fm.Fund_Currency, fm.Hedging_Policy,
               ROUND(fx.value*100, 2)        AS fx_contribution_pct,
               ROUND(fxa.value*100, 2)        AS fx_contribution_ann_pct,
               ROUND(fxv.value*100, 2)        AS fx_volatility_pct,
               ROUND(ret.value*100, 2)        AS return_total_pct,
               CAST(srri.value AS INTEGER)     AS srri
        FROM fund_metrics fx
        JOIN fund_master fm    ON fm.ISIN=fx.isin
        LEFT JOIN fund_metrics fxa  ON fxa.isin=fx.isin  AND fxa.metric='fx_contribution_ann'
                                   AND fxa.horizon='since_inception' AND fxa.real_flag=0
        LEFT JOIN fund_metrics fxv  ON fxv.isin=fx.isin  AND fxv.metric='fx_volatility_ann'
                                   AND fxv.horizon='since_inception' AND fxv.real_flag=0
        LEFT JOIN fund_metrics ret  ON ret.isin=fx.isin  AND ret.metric='return_ann'
                                   AND ret.horizon='since_inception' AND ret.real_flag=0
        LEFT JOIN fund_metrics srri ON srri.isin=fx.isin AND srri.metric='srri_nav'
                                   AND srri.horizon='since_inception' AND srri.real_flag=0
        WHERE fx.metric='fx_contribution_pct' AND fx.horizon='since_inception'
          AND fx.real_flag=0 AND fx.value IS NOT NULL
        ORDER BY ABS(fx.value) DESC
    """).fetchall()


def build_divisa(ws, conn):
    _no_gridlines(ws)
    ws.merge_cells(f"A1:{get_column_letter(10)}1")
    ws["A1"] = "FACTOR DIVISA -- contribucion del tipo de cambio al retorno total"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["ISIN", "Nombre", "Naturaleza", "Divisa fondo",
               "Cobertura", "% Retorno por divisa", "Contrib. divisa anual %",
               "Volatilidad divisa %", "Retorno total %", "SRRI"]
    _apply_header(ws, headers, row=2)

    rows = q_divisa(conn)
    for r in rows:
        row_idx = ws.max_row + 1
        for c_idx, val in enumerate(r, 1):
            cell = ws.cell(row_idx, c_idx, val)
            cell.font = _font()
        fx_pct = r[5]
        if fx_pct is not None:
            cell = ws.cell(row_idx, 6)
            v = float(fx_pct)
            if v > 50:
                cell.fill = AMBER_FILL; cell.font = AMBER_FONT
            elif v < -50:
                cell.fill = RED_FILL;  cell.font = RED_FONT
        srri_val = r[9]
        if srri_val is not None:
            bg, fg = SRRI_COLORS.get(int(srri_val), ("FFFFFF","000000"))
            ws.cell(row_idx, 10).fill = _fill(bg)
            ws.cell(row_idx, 10).font = _font(color=fg)

    ws.auto_filter.ref = f"A2:{get_column_letter(10)}{ws.max_row}"
    _freeze(ws, "A3")
    _autofit(ws)

# ============================================================
# Hoja 11 — Retornos por régimen macro
# ============================================================

_REGIMES_ORDER = [
    ("expansion",              "Expansión"),
    ("recalentamiento",        "Recalentamiento"),
    ("recalentamiento_tardio", "Recalentamiento Tardío"),
    ("estanflacion",           "Estanflación"),
    ("contraccion",            "Contracción"),
    ("shock_energetico",       "Shock Energético"),
    ("crisis_financiera",      "Crisis Financiera"),
]


def q_regime_returns(conn) -> list:
    selects = []
    joins   = []
    for suffix, _ in _REGIMES_ORDER:
        a = suffix[:5]
        selects += [
            f"ROUND(r_{a}.value*100,2)    AS ret_{suffix}",
            f"ROUND(s_{a}.value,3)         AS shr_{suffix}",
            f"ROUND(v_{a}.value*100,2)     AS vol_{suffix}",
            f"CAST(n_{a}.value AS INTEGER) AS obs_{suffix}",
        ]
        joins.append(f"""
        LEFT JOIN fund_metrics r_{a} ON r_{a}.isin=fm.ISIN
            AND r_{a}.metric='return_ann_{suffix}'
            AND r_{a}.horizon='since_inception' AND r_{a}.real_flag=0
        LEFT JOIN fund_metrics s_{a} ON s_{a}.isin=fm.ISIN
            AND s_{a}.metric='sharpe_{suffix}'
            AND s_{a}.horizon='since_inception' AND s_{a}.real_flag=0
        LEFT JOIN fund_metrics v_{a} ON v_{a}.isin=fm.ISIN
            AND v_{a}.metric='vol_ann_{suffix}'
            AND v_{a}.horizon='since_inception' AND v_{a}.real_flag=0
        LEFT JOIN fund_metrics n_{a} ON n_{a}.isin=fm.ISIN
            AND n_{a}.metric='n_obs_{suffix}'
            AND n_{a}.horizon='since_inception' AND n_{a}.real_flag=0""")

    or_clause = " OR ".join(
        f"n_{s[:5]}.value IS NOT NULL" for s, _ in _REGIMES_ORDER
    )
    sql = f"""
        SELECT fm.ISIN, fm.Fund_Name, fm.Fund_Nature, fm.Management_Company,
               {", ".join(selects)}
        FROM fund_master fm
        {"".join(joins)}
        WHERE ({or_clause})
        ORDER BY fm.Fund_Nature, fm.Fund_Name
    """
    return conn.execute(sql).fetchall()


def build_regime_returns(ws, conn):
    _no_gridlines(ws)
    n_regime_cols = len(_REGIMES_ORDER) * 4
    N_COLS = 4 + n_regime_cols

    ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
    ws["A1"] = "RENTABILIDAD POR RÉGIMEN MACRO — pipeline v10 (mín. 12 meses en régimen)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    col = 5
    for _, label in _REGIMES_ORDER:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+3)
        cell = ws.cell(2, col, label)
        cell.font = _font(bold=True, color="FFFFFF")
        cell.fill = TITLE_FILL
        cell.alignment = Alignment(horizontal="center")
        col += 4

    base_hdrs = ["ISIN", "Nombre", "Naturaleza", "Gestora"]
    regime_hdrs = []
    for suffix, _ in _REGIMES_ORDER:
        regime_hdrs += [f"Ret% {suffix[:6]}", f"Sharpe {suffix[:6]}",
                        f"Vol% {suffix[:6]}", f"N obs {suffix[:6]}"]
    _apply_header(ws, base_hdrs + regime_hdrs, row=3)

    rows = q_regime_returns(conn)
    for r in rows:
        row_idx = ws.max_row + 1
        for c_idx, val in enumerate(r, 1):
            cell = ws.cell(row_idx, c_idx, val)
            cell.font = _font()
            offset = c_idx - 5
            if offset >= 0 and offset % 4 == 0 and val is not None:
                try:
                    v = float(val)
                    if v >= 5.0:
                        cell.fill = GREEN_FILL; cell.font = GREEN_FONT
                    elif v < 0.0:
                        cell.fill = RED_FILL;   cell.font = RED_FONT
                except (TypeError, ValueError):
                    pass

    ws.auto_filter.ref = f"A3:{get_column_letter(N_COLS)}{ws.max_row}"
    _freeze(ws, "A4")
    _autofit(ws)



SHEETS = [
    ("1_Estado",        build_estado,       False),
    ("2_SRRI",          build_srri,         False),
    ("3_Rentabilidad",  build_rentabilidad, False),
    ("4_Riesgo",        build_riesgo,       False),
    ("5_Consistencia",  build_consistencia, False),
    ("6_Crisis",        build_crisis,       False),
    ("7_Candidatos",    build_candidatos,   False),
    ("8_Macro_Betas",   build_macro,        False),
    ("9_Persistencia",  build_persistencia, False),
    ("10_Divisa",       build_divisa,       False),
    ("11_Regimen_Ret",  build_regime_returns, False),
]


def export(output_dir: Path, min_fondos: int = 100) -> Path:
    conn = get_connection()

    fondos = conn.execute(
        "SELECT COUNT(DISTINCT isin) FROM fund_metrics").fetchone()[0]
    if fondos < min_fondos:
        print(f"AVISO: Solo {fondos} fondos procesados (minimo esperado: {min_fondos}).")
        print("El pipeline puede no haber terminado. Usa --min-fondos 0 para forzar.")

    ts      = datetime.now()
    ts_str  = ts.strftime("%Y-%m-%d %H:%M:%S")
    ts_file = ts.strftime("%Y%m%d_%H%M%S")

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"p2_metricas_{ts_file}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    for sheet_name, builder, _ in SHEETS:
        t0 = time.time()
        ws = wb.create_sheet(sheet_name)
        try:
            if sheet_name == "0_Portada":
                builder(ws, conn, ts_str)
            else:
                builder(ws, conn)
            elapsed = time.time() - t0
            print(f"  [{sheet_name}] OK ({elapsed:.1f}s)")
        except Exception as e:
            ws["A1"] = f"ERROR al generar esta hoja: {e}"
            ws["A1"].font = _font(bold=True, color="9C0006")
            print(f"  [{sheet_name}] ERROR: {e}")

    conn.close()
    wb.save(str(out_path))
    print(f"\nExcel generado: {out_path}")
    return out_path


# ============================================================
# Entry point
# ============================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Exporta metricas P2 a Excel")
    parser.add_argument(
        "--output", default=str(METRICS_DIR),
        help="Directorio de salida (default: out/metrics/ bajo raiz del proyecto)")
    parser.add_argument(
        "--min-fondos", type=int, default=2500,
        help="Minimo de fondos procesados para continuar (default: 2500)")
    args = parser.parse_args()

    print(f"\nExportando metricas P2...")
    print(f"  BD:      {DB_PATH}")
    print(f"  Salida:  {args.output}\\")
    print()

    export(
        output_dir=Path(args.output),
        min_fondos=args.min_fondos,
    )
