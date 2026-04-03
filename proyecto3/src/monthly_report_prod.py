# proyecto3/src/monthly_report.py
# -*- coding: utf-8 -*-
"""
Generador de informe mensual de la cartera.

Produce un Excel con las siguientes hojas:
  0_Portada         Resumen ejecutivo con regimen, semaforo y cartera
  1_Cartera         Detalle de fondos con pesos y metricas clave
  2_Regimen         Clasificacion macro actual e historica
  3_Backtesting     Resumen de resultados del backtesting
  4_Rotacion        Recomendaciones de rotacion si procede

Uso:
    from proyecto3.src.monthly_report import generate_report
    generate_report(conn, output_dir="c:/data/fondos/reports")
"""

import sqlite3
import pandas as pd
from pathlib import Path
from datetime import date
import sys

_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_ROOT))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl no instalado. Ejecuta: pip install openpyxl")

from proyecto3.src.regime_classifier import RegimeClassifier
from proyecto3.src.backtesting import Backtester
from shared.config import REPORTS_DIR as _DEFAULT_REPORTS_DIR


# ============================================================
# Estilos
# ============================================================

def _font(bold=False, size=11, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _center():
    return Alignment(horizontal="center", vertical="center")

TITLE_FONT  = _font(bold=True, size=13, color="FFFFFF")
TITLE_FILL  = _fill("1D3557")
HEADER_FONT = _font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = _fill("457B9D")

REGIME_COLORS = {
    "Expansion":              ("D4EDDA", "155724"),
    "Recalentamiento":        ("FFF3CD", "856404"),
    "Recalentamiento_Tardio": ("FFE5B4", "6B3A00"),
    "Estanflacion":           ("F8D7DA", "721C24"),
    "Contraccion":            ("D1ECF1", "0C5460"),
    "Shock_Energetico":       ("F8D7DA", "721C24"),
}

SEMAFORO_COLORS = {
    "Verde": ("D4EDDA", "155724"),
    "Ambar": ("FFF3CD", "856404"),
    "Rojo":  ("F8D7DA", "721C24"),
}


def _apply_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row, col, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = _center()


def _autofit(ws, min_width=8, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ============================================================
# Hojas del informe
# ============================================================

def _build_portada(ws, conn, clf, scenario_id="shock_energia_2026Q1"):
    """Hoja 0 — Portada con resumen ejecutivo."""
    ws.sheet_view.showGridLines = False

    # Titulo
    ws.merge_cells("A1:F1")
    ws["A1"] = f"INFORME MENSUAL DE CARTERA  |  {date.today().strftime('%B %Y').upper()}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 30

    # Regimen actual
    r = clf.classify_current()
    sem = clf.semaforo()
    reg_bg, reg_fg = REGIME_COLORS.get(r.regime, ("FFFFFF", "000000"))
    sem_bg, sem_fg = SEMAFORO_COLORS.get(sem.color, ("FFFFFF", "000000"))

    datos_regimen = [
        ("", ""),
        ("REGIMEN MACRO", r.regime),
        ("SEMAFORO", sem.color),
        ("IPC medio EU+ES", f"{r.ipc_yoy_avg*100:.1f}%" if r.ipc_yoy_avg else "N/D"),
        ("CLI Eurozona", f"{r.cli_eu:.1f}" if r.cli_eu else "N/D"),
        ("Tipo deposito BCE", f"{r.rate_deposit:.2f}%" if r.rate_deposit else "N/D"),
        ("Petroleo YoY", f"{r.oil_yoy*100:+.1f}%" if r.oil_yoy else "N/D"),
        ("", ""),
        ("PESOS DE CARTERA", ""),
        ("Defensiva", f"{r.weight_defensive:.0%}"),
        ("Equilibrada", f"{r.weight_balanced:.0%}"),
        ("Dinamica", f"{r.weight_dynamic:.0%}"),
    ]

    for i, (label, value) in enumerate(datos_regimen, 3):
        ws.cell(i, 1, label).font = _font(bold=True, size=10)
        cell_val = ws.cell(i, 2, value)
        cell_val.font = _font(size=10)
        if label == "REGIMEN MACRO":
            cell_val.fill = _fill(reg_bg)
            cell_val.font = _font(bold=True, size=10, color=reg_fg)
        elif label == "SEMAFORO":
            cell_val.fill = _fill(sem_bg)
            cell_val.font = _font(bold=True, size=10, color=sem_fg)

    # Senales del semaforo
    if sem.signals:
        row = 3 + len(datos_regimen) + 1
        ws.cell(row, 1, "SENALES:").font = _font(bold=True, size=10)
        for j, signal in enumerate(sem.signals):
            ws.cell(row + j + 1, 1, f"  ! {signal}").font = _font(size=10, color="721C24")

    _autofit(ws)


def _build_cartera(ws, conn, scenario_id="shock_energia_2026Q1"):
    """Hoja 1 — Detalle de fondos de la cartera actual."""
    ws.sheet_view.showGridLines = False

    ws["A1"] = f"CARTERA ACTUAL  |  Escenario: {scenario_id}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = _center()

    headers = ["Sub-cartera", "ISIN", "Nombre", "Naturaleza", "Gestora",
               "Peso maestro %", "Score", "Retorno real %", "Max DD %"]
    _apply_header(ws, headers, row=2)
    ws.merge_cells(f"A1:{get_column_letter(9)}1")

    rows = conn.execute("""
        SELECT pw.block, pw.isin, fm.Fund_Name, fm.Fund_Nature,
               fm.Management_Company,
               ROUND(pw.weight * 100, 1) as peso_master,
               ROUND(fs.score_total, 4) as score,
               ROUND(ret.value * 100, 2) as return_real,
               ROUND(dd.value * 100, 2) as max_dd
        FROM portfolio_weights pw
        JOIN fund_master fm ON fm.ISIN = pw.isin
        LEFT JOIN fund_scores fs ON fs.isin = pw.isin AND fs.block = pw.block
                                 AND fs.score_version = 'v1'
        LEFT JOIN fund_metrics ret ON ret.isin = pw.isin
                                   AND ret.metric = 'return_ann'
                                   AND ret.horizon = 'since_inception'
                                   AND ret.real_flag = 1
        LEFT JOIN fund_metrics dd  ON dd.isin = pw.isin
                                   AND dd.metric = 'max_drawdown'
                                   AND dd.horizon = 'since_inception'
                                   AND dd.real_flag = 0
        WHERE pw.scenario_id = ?
        ORDER BY pw.block, pw.weight DESC
    """, (scenario_id,)).fetchall()

    sub_colors = {
        "Defensiva":   "D1ECF1",
        "Equilibrada": "FFF3CD",
        "Dinamica":    "F8D7DA",
    }

    for r in rows:
        row_idx = ws.max_row + 1
        for c_idx, val in enumerate(r, 1):
            cell = ws.cell(row_idx, c_idx, val)
            cell.font = _font(size=10)
            bg = sub_colors.get(r[0], "FFFFFF")
            cell.fill = _fill(bg)

    ws.auto_filter.ref = f"A2:{get_column_letter(10)}{ws.max_row}"
    ws.freeze_panes = "A3"
    _autofit(ws)


def _build_regimen(ws, clf):
    """Hoja 2 — Clasificacion macro actual e historica."""
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"A1:{get_column_letter(8)}1")
    ws["A1"] = "CLASIFICACION DE REGIMEN MACRO"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = _center()

    # Distribucion historica
    ws.cell(3, 1, "DISTRIBUCION HISTORICA 2000-2026").font = _font(bold=True, size=11)
    summary = clf.regime_summary()

    headers = ["Regimen", "Meses", "% Tiempo"]
    _apply_header(ws, headers, row=4)
    for i, (regime, row) in enumerate(summary.iterrows()):
        r_idx = 5 + i
        ws.cell(r_idx, 1, regime).font = _font(size=10)
        ws.cell(r_idx, 2, int(row["meses"])).font = _font(size=10)
        ws.cell(r_idx, 3, f"{row['pct']:.1f}%").font = _font(size=10)
        bg, fg = REGIME_COLORS.get(regime, ("FFFFFF", "000000"))
        for c in range(1, 4):
            ws.cell(r_idx, c).fill = _fill(bg)

    # Ultimos 24 meses
    hist = clf.classify_historical()
    if not hist.empty:
        ws.cell(ws.max_row + 2, 1,
                "CLASIFICACION MENSUAL (ultimos 24 meses)").font = _font(bold=True, size=11)
        headers2 = ["Fecha", "Regimen", "Defensiva", "Equilibrada", "Dinamica",
                    "IPC avg", "CLI EU", "Tipo", "Petroleo YoY"]
        _apply_header(ws, headers2, row=ws.max_row + 1)

        recent = hist.iloc[-24:]
        for date, row in recent.iterrows():
            r_idx = ws.max_row + 1
            vals = [
                date.strftime("%Y-%m"),
                row["regime"],
                f"{row['weight_defensive']:.0%}",
                f"{row['weight_balanced']:.0%}",
                f"{row['weight_dynamic']:.0%}",
                f"{row['ipc_yoy_avg']*100:.1f}%" if row.get('ipc_yoy_avg') else "N/D",
                f"{row['cli_eu']:.1f}" if row.get('cli_eu') else "N/D",
                f"{row['rate_deposit']:.2f}%" if row.get('rate_deposit') else "N/D",
                f"{row['oil_yoy']*100:+.1f}%" if row.get('oil_yoy') else "N/D",
            ]
            for c_idx, val in enumerate(vals, 1):
                cell = ws.cell(r_idx, c_idx, val)
                cell.font = _font(size=9)
            bg, _ = REGIME_COLORS.get(row["regime"], ("FFFFFF", "000000"))
            ws.cell(r_idx, 2).fill = _fill(bg)

    _autofit(ws)


def _build_backtesting(ws, conn):
    """Hoja 3 — Resumen del backtesting."""
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"A1:{get_column_letter(7)}1")
    ws["A1"] = "BACKTESTING -- Validacion historica del modelo (2005-2026)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = _center()

    bt = Backtester(conn)
    results = bt.run(start_date="2005-01-01")

    ws.cell(3, 1, "RESUMEN POR REGIMEN (horizonte 12 meses)").font = _font(bold=True, size=11)
    headers = ["Regimen", "Meses", "Cartera 12m %",
               "Benchmark 12m %", "Exceso pp", "Hit ratio %"]
    _apply_header(ws, headers, row=4)

    for regime in results["regime"].unique():
        sub = results[results["regime"] == regime]
        valid = sub[["ret_12m", "bench_12m", "excess_12m"]].dropna()
        if valid.empty:
            continue
        r_idx = ws.max_row + 1
        vals = [
            regime,
            len(sub),
            f"{valid['ret_12m'].mean()*100:+.1f}%",
            f"{valid['bench_12m'].mean()*100:+.1f}%",
            f"{valid['excess_12m'].mean()*100:+.1f}pp",
            f"{(valid['excess_12m'] > 0).mean()*100:.0f}%",
        ]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.font = _font(size=10)
        bg, _ = REGIME_COLORS.get(regime, ("FFFFFF", "000000"))
        ws.cell(r_idx, 1).fill = _fill(bg)

    # Global
    r_idx = ws.max_row + 2
    ws.cell(r_idx, 1, "GLOBAL").font = _font(bold=True, size=10)
    valid_all = results[["ret_12m", "bench_12m", "excess_12m"]].dropna()
    if not valid_all.empty:
        ws.cell(r_idx, 2, len(results))
        ws.cell(r_idx, 3, f"{valid_all['ret_12m'].mean()*100:+.1f}%").font = _font(bold=True, size=10)
        ws.cell(r_idx, 4, f"{valid_all['bench_12m'].mean()*100:+.1f}%").font = _font(bold=True, size=10)
        ws.cell(r_idx, 5, f"{valid_all['excess_12m'].mean()*100:+.1f}pp").font = _font(bold=True, size=10, color="155724")
        ws.cell(r_idx, 6, f"{(valid_all['excess_12m'] > 0).mean()*100:.0f}%").font = _font(bold=True, size=10)

    ws.cell(r_idx + 2, 1,
            "Nota: backtesting simple con look-ahead bias. "
            "Las metricas usan datos completos hasta la fecha de calculo.").font = _font(size=9, color="888780")

    _autofit(ws)


# ============================================================
# Generador principal
# ============================================================

def generate_report(
    conn: sqlite3.Connection,
    output_dir: str | None = None,
    scenario_id: str = "shock_energia_2026Q1",
    score_version: str = "v1",
) -> Path:
    """
    Genera el informe mensual completo en Excel.

    Devuelve la ruta del fichero generado.
    """
    today      = date.today().strftime("%Y%m%d")
    filename   = f"informe_cartera_{today}.xlsx"
    if output_dir is None:
        output_dir = str(_DEFAULT_REPORTS_DIR)
    output_path = Path(output_dir) / filename
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    clf = RegimeClassifier(conn)

    wb = Workbook()
    wb.remove(wb.active)  # eliminar hoja por defecto

    print("Generando informe mensual...")

    ws0 = wb.create_sheet("0_Portada")
    _build_portada(ws0, conn, clf, scenario_id)
    print("  0_Portada OK")

    ws1 = wb.create_sheet("1_Cartera")
    _build_cartera(ws1, conn, scenario_id)
    print("  1_Cartera OK")

    ws2 = wb.create_sheet("2_Regimen")
    _build_regimen(ws2, clf)
    print("  2_Regimen OK")

    ws3 = wb.create_sheet("3_Backtesting")
    _build_backtesting(ws3, conn)
    print("  3_Backtesting OK")

    wb.save(output_path)
    print(f"\nInforme generado: {output_path}")
    return output_path
