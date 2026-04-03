# -*- coding: utf-8 -*-
"""
core/io.py

IO documental canónico:
- indexa enlaces KIID en el Excel maestro (una sola pasada)
- descarga PDFs con retries
- extrae texto (pdfplumber + OCR opcional)
- devuelve texto y metadata estructurada (incluye KIID_PDF_BYTES cuando hay texto)
"""

from typing import Optional, Dict, List, Tuple
from io import BytesIO
import hashlib
import datetime
import gc

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

import pdfplumber
try:
    import pytesseract
    _HAS_TESSERACT = True
except Exception:
    _HAS_TESSERACT = False

from openpyxl import load_workbook

# Config
HEADERS = {"User-Agent": "Mozilla/5.0"}
REQUEST_TIMEOUT  = 15     # timeout por intento en segundos
REQUEST_RETRIES  = 3     # reintentos tras fallo (4 intentos totales)
# Demoras entre reintentos: backoff_factor=1 → 1s, 2s, 4s (formula: factor × 2^(n-1))

OCR_ENABLED = True
OCR_LANG = "spa+eng"
OCR_DPI = 300


MAX_PDF_PAGES = 3        # KIID = 2–3 páginas

KIID_CACHE_DAYS = 180    # días antes de re-verificar el PDF (Opción A)

MAX_PDF_MB = 20
MAX_PDF_BYTES = MAX_PDF_MB * 1024 * 1024


# Cache ligero: excel_path -> { ISIN: [url, ...] }
_ISIN_LINKS_INDEX: Dict[str, Dict[str, List[str]]] = {}


def _requests_session() -> requests.Session:
    s = requests.Session()
    retries = Retry(
        total            = REQUEST_RETRIES,
        backoff_factor   = 1,                          # esperas: 1s, 2s, 4s
        status_forcelist = (500, 502, 503, 504),       # errores de servidor
        raise_on_status  = False,
    )
    s.mount("https://", HTTPAdapter(max_retries=retries))
    s.mount("http://", HTTPAdapter(max_retries=retries))
    s.headers.update(HEADERS)
    return s


def extract_text_from_pdf_bytes(
    pdf_bytes: bytes,
    ocr_enabled: bool = OCR_ENABLED,
    ocr_lang: str = OCR_LANG,
    ocr_dpi: int = OCR_DPI,
) -> str:
    text_parts = []

    with BytesIO(pdf_bytes) as b:
        with pdfplumber.open(b) as pdf:
            for i, page in enumerate(pdf.pages):
                if i >= MAX_PDF_PAGES:
                    break

                try:
                    t = page.extract_text()
                except Exception:
                    t = None

                if t and t.strip():
                    text_parts.append(t)
                    continue

                if ocr_enabled and _HAS_TESSERACT:
                    try:
                        img = page.to_image(resolution=ocr_dpi).original
                        text_parts.append(
                            pytesseract.image_to_string(img, lang=ocr_lang)
                        )
                    except Exception:
                        continue

    return "\n".join(text_parts)



def pdf_sha256(pdf_bytes: bytes) -> str:
    return hashlib.sha256(pdf_bytes).hexdigest()



# ============================================================
# TRANSICIÓN AUTOMÁTICA A FORCE_REFRESH POR ANTIGÜEDAD
# ============================================================

def mark_stale_for_refresh(
    conn,
    max_age_days: int = KIID_CACHE_DAYS,
    max_funds: int = 100,
) -> int:
    """
    Marca como FORCE_REFRESH los fondos cuyo KIID no se ha re-descargado
    en más de max_age_days días.

    Parámetros:
        conn         — conexión SQLite activa
        max_age_days — antigüedad máxima en días (por defecto KIID_CACHE_DAYS=180)
        max_funds    — máximo de fondos a marcar por ejecución (evita avalanchas)

    Devuelve el número de fondos marcados.

    Llamar UNA VEZ al inicio del ciclo, antes de lanzar los bloques:
        from core.io import mark_stale_for_refresh
        n = mark_stale_for_refresh(conn, max_age_days=180, max_funds=50)
        print(f"[INFO] {n} fondos marcados FORCE_REFRESH por antigüedad")

    Con max_funds=50 y un universo de 3.204 fondos, el refresh completo
    se distribuye en ~64 ciclos (180 días / 50 fondos/ciclo ≈ 3,6 fondos/día
    de media), sin generar avalanchas de requests al servidor.
    """
    cutoff = (
        datetime.datetime.utcnow() - datetime.timedelta(days=max_age_days)
    ).isoformat(timespec="seconds")

    # Seleccionar los más antiguos primero (FIFO de antigüedad)
    sql = """
        UPDATE fund_kiid_metadata
        SET    KIID_Status = 'FORCE_REFRESH'
        WHERE  ISIN IN (
            SELECT ISIN
            FROM   fund_kiid_metadata
            WHERE  KIID_Status   IN ('OK', 'CACHED')
            AND    KIID_Class     = 1
            AND    (KIID_Downloaded_At IS NULL
                    OR KIID_Downloaded_At < ?)
            ORDER  BY KIID_Downloaded_At ASC NULLS FIRST
            LIMIT  ?
        )
        AND    KIID_Class = 1;
    """
    cursor = conn.execute(sql, (cutoff, max_funds))
    conn.commit()
    return cursor.rowcount


def _build_isin_links_index(excel_path: str) -> Dict[str, List[str]]:
    """
    Abre el workbook (modo normal) y construye un dict ISIN -> [URL...].
    Guarda el índice en _ISIN_LINKS_INDEX[excel_path].
    """
    excel_path = str(excel_path)
    if excel_path in _ISIN_LINKS_INDEX:
        return _ISIN_LINKS_INDEX[excel_path]

    index: Dict[str, List[str]] = {}
    wb = load_workbook(excel_path, data_only=True, read_only=False)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2):
                try:
                    if len(row) < 3:
                        continue
                    cell_isin = row[2]
                    if cell_isin is None or cell_isin.value is None:
                        continue
                    isin = str(cell_isin.value).strip()
                    if not isin:
                        continue

                    urls = index.setdefault(isin, [])
                    cell_link = row[3] if len(row) > 3 else None

                    if cell_link is not None:
                        link_obj = getattr(cell_link, "hyperlink", None)
                        if link_obj and getattr(link_obj, "target", None):
                            url = getattr(link_obj, "target")
                            if isinstance(url, str) and url.strip():
                                urls.append(url.strip())
                                continue
                        val = cell_link.value
                        if isinstance(val, str) and "http" in val.lower():
                            urls.append(val.strip())
                            continue
                except Exception:
                    continue
    finally:
        try:
            wb.close()
        except Exception:
            pass

    # normalizar: dedupe y ordenar por heurística (determinismo)
    for k, v in list(index.items()):
        seen = []
        for u in v:
            if u not in seen:
                seen.append(u)
        def _score(u):
            score = 0
            if isinstance(u, str) and u.lower().endswith(".pdf"):
                score -= 10
            if isinstance(u, str) and "kiid" in u.lower():
                score -= 5
            return score
        seen.sort(key=lambda x: (_score(x), x))
        index[k] = seen

    _ISIN_LINKS_INDEX[excel_path] = index
    gc.collect()
    return index


def find_kiid_links_from_excel(isin: str, excel_path: str) -> List[str]:
    """
    Devuelve lista (posible vacía) de URLs para un ISIN usando el índice.
    """
    index = _build_isin_links_index(excel_path)
    return index.get(isin, [])


def download_pdf(url: str, session: Optional[requests.Session] = None) -> bytes:
    session = session or _requests_session()
    resp = session.get(url, timeout=REQUEST_TIMEOUT, stream=True)
    resp.raise_for_status()

    content = bytearray()
    for chunk in resp.iter_content(chunk_size=8192):
        content.extend(chunk)
        if len(content) > MAX_PDF_BYTES:
            raise ValueError("pdf_too_large")

    return bytes(content)




def get_kiid_for_isin(
    isin: str,
    excel_master_path: Optional[str],
    conn=None,
) -> Tuple[Optional[str], Dict]:
    """
    Intentar localizar y extraer el texto del KIID para un ISIN.

    Devuelve:
        (kiid_text o None, metadata dict)

    metadata incluye:
        - KIID_Status  (estado documental: OK / CACHED / DOWNLOAD_ERROR / ...)
        - KIID_Error   (detalle técnico)
        - KIID_PDF_BYTES (solo si KIID_Status == OK — descarga real)

    Si conn es proporcionado, comprueba primero si ya existe Raw_KIID_Text
    en fund_kiid_metadata (KIID_Status=OK). En ese caso devuelve el texto
    cacheado sin re-descargar el PDF.
    """

    meta = {
        "ISIN": isin,
        "KIID_URL": None,
        "KIID_PDF_Hash": None,
        "KIID_Downloaded_At": None,
        "KIID_Status": None,
        "KIID_Error": None,
    }

    # -------------------------------------------------
    # Caché A+B:
    #   A) Si el KIID fue descargado hace < KIID_CACHE_DAYS → usar texto BD
    #   B) Si han pasado más días → descargar PDF y comparar hash:
    #      - Hash igual  → reusar texto BD (evita OCR + SRRI visual)
    #      - Hash distinto → procesar PDF completo
    # -------------------------------------------------
    _cached_text = _cached_hash = _cached_url = None
    if conn is not None:
        try:
            cached = conn.execute(
                """SELECT Raw_KIID_Text, KIID_PDF_Hash, KIID_URL,
                          KIID_Downloaded_At
                   FROM fund_kiid_metadata
                   WHERE ISIN = ? AND KIID_Class = 1
                     AND KIID_Status IN ('OK', 'CACHED')
                     AND Raw_KIID_Text IS NOT NULL""",
                (isin,)
            ).fetchone()
            if cached and cached[0] and cached[0].strip():
                _cached_text = cached[0]
                _cached_hash = cached[1]
                _cached_url  = cached[2]
                downloaded_at = cached[3]

                # Devolver texto de BD sin descarga.
                # KIID_Downloaded_At se preserva con su valor original de la BD:
                # es la fecha en que el PDF fue descargado físicamente por última vez.
                # No se sobreescribe en lecturas de caché — solo en descargas reales.
                # La re-descarga se activa únicamente mediante FORCE_REFRESH explícito.
                meta.update({
                    "KIID_URL":            _cached_url,
                    "KIID_PDF_Hash":       _cached_hash,
                    "KIID_Status":         "CACHED",
                    "KIID_Downloaded_At":  downloaded_at,  # preservar fecha original
                    "KIID_Error":          None,
                    "KIID_Class":          1,
                })
                return _cached_text, meta
        except Exception:
            _cached_text = _cached_hash = _cached_url = None  # fallo → descarga normal

    # -------------------------------------------------
    # Validación inicial
    # -------------------------------------------------
    if not excel_master_path:
        meta["KIID_Status"] = "DOWNLOAD_ERROR"
        meta["KIID_Error"] = "no_excel_master_path"
        return None, meta

    try:
        links = find_kiid_links_from_excel(isin, excel_master_path)
    except Exception as e:
        meta["KIID_Status"] = "DOWNLOAD_ERROR"
        meta["KIID_Error"] = f"excel_index_error: {e}"
        return None, meta

    if not links:
        meta["KIID_Status"] = "NO_KIID"
        meta["KIID_Error"] = "no_links_found"
        return None, meta

    # -------------------------------------------------
    # Intentar cada enlace
    # -------------------------------------------------
    session = _requests_session()

    for url in links:
        try:
            pdf_bytes = download_pdf(url, session)

            # Opción B: comparar hash — solo llega aquí para FORCE_REFRESH
            # (la Opción A devuelve siempre para fondos con texto en BD).
            new_hash = pdf_sha256(pdf_bytes)
            if _cached_text and _cached_hash and new_hash == _cached_hash:
                # PDF idéntico — reusar texto BD, evitar OCR + SRRI visual.
                # KIID_Status = "OK": el PDF fue descargado físicamente con éxito.
                # KIID_Downloaded_At = ahora: registrar la fecha real de descarga.
                # El escritor usará COALESCE pero este valor no es NULL → se persiste.
                # La próxima ejecución verá una fecha reciente → Opción A → sin descarga.
                meta.update({
                    "KIID_URL":            url,
                    "KIID_PDF_Hash":       new_hash,
                    "KIID_Downloaded_At":  datetime.datetime.utcnow().isoformat(timespec="seconds"),
                    "KIID_Status":         "OK",   # descarga real exitosa, aunque texto sin cambio
                    "KIID_Error":          None,
                    "KIID_Class":          1,
                })
                return _cached_text, meta

            # PDF nuevo o modificado — procesar completo
            try:
                kiid_text = extract_text_from_pdf_bytes(pdf_bytes)
            except Exception as e:
                meta["KIID_Status"] = "DOWNLOAD_ERROR"
                meta["KIID_Error"] = f"text_extraction_error: {e}"
                continue

            if not kiid_text or not kiid_text.strip():
                meta["KIID_Status"] = "EMPTY_TEXT"
                meta["KIID_Error"] = f"empty_text_from_{url}"
                continue

            # -------------------------------------------------
            # KIID válido — procesado completo
            # -------------------------------------------------
            meta.update({
                "KIID_URL": url,
                "KIID_PDF_Hash": new_hash,
                "KIID_Downloaded_At": datetime.datetime.utcnow().isoformat(timespec="seconds"),
                "KIID_PDF_BYTES": pdf_bytes,   # temporal (pipeline)
                "KIID_Status": "OK",
                "KIID_Error": None,
            })

            return kiid_text, meta

        except ValueError as e:
            # Casos controlados (ej. pdf_too_large)
            if "pdf_too_large" in str(e):
                meta["KIID_Status"] = "PDF_TOO_LARGE"
                meta["KIID_Error"] = str(e)
            else:
                meta["KIID_Status"] = "DOWNLOAD_ERROR"
                meta["KIID_Error"] = str(e)
            continue

        except Exception as e:
            meta["KIID_Status"] = "DOWNLOAD_ERROR"
            meta["KIID_Error"] = str(e)
            continue

    # -------------------------------------------------
    # Ningún enlace válido
    # -------------------------------------------------
    if meta["KIID_Status"] is None:
       meta["KIID_Status"] = "NO_KIID"
       meta["KIID_Error"] = "no_valid_kiid_found"


    return None, meta
